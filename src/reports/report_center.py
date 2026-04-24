# src/reports/report_center.py
from __future__ import annotations
import tkinter as tk
from tkinter import ttk, messagebox
from typing import List, Optional, Tuple
from pathlib import Path
import configparser
import datetime as dt
import webbrowser
from decimal import Decimal
from sqlalchemy import func

from PIL import Image, ImageDraw, ImageFont

from src.data.database import get_session
from src.data.models import (
    Product, Supplier, Customer,
    Purchase, PurchaseDetail, PurchasePayment, Reception,
    Sale, SaleDetail, SaleServiceDetail,
)
from src.gui.utils.order_helpers import format_currency
from src.gui.widgets.grid_table import GridTable

# Stock real: servicio + exportar/imprimir
from src.reports.inventory_reports import (
    InventoryFilter,
    InventoryReportService,
    generate_inventory_xlsx,
    print_inventory_report,
)
from src.gui.printer_select_dialog import PrinterSelectDialog
from src.utils.printers import get_document_printer


# ----------------------------- Utilidades ----------------------------- #
def _parse_date(s: str) -> Optional[dt.datetime]:
    """Convierte YYYY-MM-DD en datetime al inicio del día. None si vacío/incorrecto."""
    s = (s or "").strip()
    if not s:
        return None
    try:
        y, m, d = s.split("-")
        return dt.datetime(int(y), int(m), int(d), 0, 0, 0)
    except Exception:
        return None


def _range_to_datetimes(d_from: Optional[dt.datetime], d_to: Optional[dt.datetime]) -> Tuple[Optional[dt.datetime], Optional[dt.datetime]]:
    """Si hay fecha hasta, muévela a 23:59:59 para incluirla completa."""
    if d_to is not None:
        d_to = d_to + dt.timedelta(hours=23, minutes=59, seconds=59)
    return d_from, d_to


def _fmt_money(value) -> str:
    return format_currency(value or 0)


def _fmt_date(val: Optional[dt.datetime], with_time: bool = False) -> str:
    if not val:
        return ""
    return val.strftime("%Y-%m-%d %H:%M") if with_time else val.strftime("%Y-%m-%d")


def _parse_number(s: str) -> Optional[float]:
    try:
        return float((s or "").replace(".", "").replace(",", ".")) if s else None
    except Exception:
        return None


def _downloads_dir() -> Path:
    home = Path.home()
    for cand in ("Downloads", "Descargas", "downloads", "DESCARGAS"):
        path = home / cand
        if path.exists():
            return path
    return home


def _read_company_cfg() -> dict[str, str]:
    data = {
        "name": "Inventario App",
        "rut": "",
        "address": "",
        "phone": "",
        "email": "",
        "logo": "",
    }
    for cfg_path in (Path("config/settings.ini"), Path("config/company.ini")):
        if not cfg_path.exists():
            continue
        cfg = configparser.ConfigParser()
        cfg.read(cfg_path, encoding="utf-8")
        section = cfg["company"] if cfg.has_section("company") else cfg["DEFAULT"]
        data["name"] = section.get("name", section.get("nombre", data["name"])) or data["name"]
        data["rut"] = section.get("rut", data["rut"]) or data["rut"]
        data["address"] = section.get("address", section.get("direccion", data["address"])) or data["address"]
        data["phone"] = section.get("phone", section.get("telefono", data["phone"])) or data["phone"]
        data["email"] = section.get("email", data["email"]) or data["email"]
        data["logo"] = section.get("logo", section.get("logo_path", data["logo"])) or data["logo"]
        break
    return data


# --------------------------- Vista principal -------------------------- #
class ReportCenter(ttk.Frame):
    """
    Centro de Informes:
      - Stock real
      - Facturas o guias (por pagar / por cobrar)
      - Ordenes de compra (resumen y detalle por proveedor/producto)
      - Ventas por periodo
      - Inversion monetaria por producto
    """

    # Definición de informes disponibles
    REPORTS = [
        ("stock_real", "Stock"),
        ("payables_docs", "Cuentas por pagar"),
        ("supplier_debts", "Deuda proveedores"),
        ("receivables_docs", "Cuentas por cobrar"),
        ("customer_debt", "Deuda clientes"),
        ("all_purchase_orders", "OC detalle"),
        ("all_sales_orders", "OV detalle"),
        ("purchase_orders", "OC resumen"),
        ("purchase_by_supplier_product", "OC por producto"),
        ("sales_period", "Ventas resumen"),
        ("price_list", "Lista precios"),
        ("investment_by_product", "Inversión stock"),
    ]

    REPORT_SUMMARIES = {
        "stock_real": "Existencias actuales por producto, con unidad, SKU, precio de compra y precio de venta.",
        "payables_docs": "Documentos de compra pendientes de pago, filtrables por proveedor, estado y periodo.",
        "supplier_debts": "Deuda agrupada por proveedor, con total comprado, pagado, saldo y vencimiento.",
        "receivables_docs": "Ventas pendientes de cobro, filtrables por cliente, estado y periodo.",
        "customer_debt": "Detalle de deuda por cliente, incluyendo factura, neto, IVA, total y fecha de pago.",
        "all_purchase_orders": "Detalle completo de OC por producto, proveedor, cantidades, recibido, subtotal y total.",
        "all_sales_orders": "Detalle completo de OV por producto o servicio, cliente, estado, subtotal y total.",
        "purchase_orders": "Resumen de OC por proveedor, estado, documento, fecha y total.",
        "purchase_by_supplier_product": "Cruce de OC por proveedor y producto para revisar cantidades, precios y subtotales.",
        "sales_period": "Resumen de ventas del periodo, con cliente, estado y total.",
        "price_list": "Lista de precios de venta por producto y código.",
        "investment_by_product": "Valor del stock disponible calculado desde precio de compra más IVA.",
    }

    SALES_STATES = ["(Todos)", "Pagado", "Pendiente"]
    RECEIVABLE_STATES = ["(Todos)", "Pendiente"]
    PURCH_STATES = ["(Todos)", "Completada", "Pendiente", "Cancelada", "Eliminada", "Por pagar", "Ingreso parcial", "Incompleta"]
    PAYABLE_STATES = ["(Todos)", "Por pagar", "Ingreso parcial", "Pendiente", "Incompleta"]

    def __init__(self, master: tk.Misc):
        super().__init__(master, padding=10)

        self.session = get_session()
        self.svc_inventory = InventoryReportService(self.session)

        # cache para exportar/imprimir
        self._current_cols: List[str] = []
        self._current_rows: List[List] = []
        self._current_report_key: str = ""
        self.var_report_summary = tk.StringVar(value="")

        # --------- Encabezado / selector ---------
        top = ttk.Frame(self); top.pack(fill="x", expand=False)
        ttk.Label(top, text="Informe:", font=("", 10, "bold")).pack(side="left")

        self.cmb_report = ttk.Combobox(
            top, state="readonly",
            values=[name for _key, name in self.REPORTS],
            width=36
        )
        self.cmb_report.current(0)
        self.cmb_report.pack(side="left", padx=8)
        self.cmb_report.bind("<<ComboboxSelected>>", lambda _e: self._on_report_changed())

        ttk.Button(top, text="Ejecutar", command=self._run_report).pack(side="right", padx=4)
        self.btn_export = ttk.Button(top, text="Exportar PDF", command=self._on_export)
        self.btn_export.pack(side="right", padx=4)
        self.btn_export_xlsx = ttk.Button(top, text="Exportar Excel", command=self._on_export_xlsx)
        self.btn_export_xlsx.pack(side="right", padx=4)

        self.lbl_report_summary = ttk.Label(
            self,
            textvariable=self.var_report_summary,
            style="HomePanelText.TLabel",
            wraplength=920,
            justify="left",
        )
        self.lbl_report_summary.pack(fill="x", pady=(8, 0))

        # --------- Filtros ---------
        filt = ttk.Labelframe(self, text="Filtros", padding=8)
        filt.pack(fill="x", expand=False, pady=(8, 4))

        ttk.Label(filt, text="Desde (YYYY-MM-DD):").grid(row=0, column=0, sticky="e", padx=4, pady=3)
        self.var_date_from = tk.StringVar(); ttk.Entry(filt, textvariable=self.var_date_from, width=14).grid(row=0, column=1, sticky="w")
        ttk.Label(filt, text="Hasta (YYYY-MM-DD):").grid(row=0, column=2, sticky="e", padx=8)
        self.var_date_to = tk.StringVar(); ttk.Entry(filt, textvariable=self.var_date_to, width=14).grid(row=0, column=3, sticky="w")

        ttk.Label(filt, text="Estado:").grid(row=0, column=4, sticky="e", padx=8)
        self.cmb_state = ttk.Combobox(filt, state="readonly", width=14, values=self.SALES_STATES)
        self.cmb_state.grid(row=0, column=5, sticky="w")
        try:
            self.cmb_state.current(0)
        except Exception:
            pass

        self.var_party_label = tk.StringVar(value="Tercero (nombre/rut):")
        self.lbl_party = ttk.Label(filt, textvariable=self.var_party_label)
        self.lbl_party.grid(row=1, column=0, sticky="e", padx=4, pady=3)
        self.var_party = tk.StringVar(); ttk.Entry(filt, textvariable=self.var_party, width=26).grid(row=1, column=1, sticky="w")
        ttk.Label(filt, text="Producto (nombre/sku):").grid(row=1, column=2, sticky="e", padx=8)
        self.var_product = tk.StringVar(); ttk.Entry(filt, textvariable=self.var_product, width=26).grid(row=1, column=3, sticky="w")

        ttk.Label(filt, text="Total min:").grid(row=1, column=4, sticky="e")
        self.var_total_min = tk.StringVar(); ttk.Entry(filt, textvariable=self.var_total_min, width=12).grid(row=1, column=5, sticky="w")
        ttk.Label(filt, text="Total max:").grid(row=1, column=6, sticky="e")
        self.var_total_max = tk.StringVar(); ttk.Entry(filt, textvariable=self.var_total_max, width=12).grid(row=1, column=7, sticky="w")

        # --------- Tabla ---------
        self.table = GridTable(self, height=16)
        self.table.pack(fill="both", expand=True, pady=(8, 0))

        self._on_report_changed()  # set estados apropiados

    # ---------------------- Helpers de filtros ---------------------- #
    def _get_date_filters(self) -> Tuple[Optional[dt.datetime], Optional[dt.datetime]]:
        d_from = _parse_date(self.var_date_from.get())
        d_to = _parse_date(self.var_date_to.get())
        return _range_to_datetimes(d_from, d_to)

    def _on_report_changed(self) -> None:
        idx = self.cmb_report.current() or 0
        key = self.REPORTS[idx][0]
        self._current_report_key = key
        self.var_report_summary.set(self.REPORT_SUMMARIES.get(key, ""))
        self._clear_report_inputs()
        self.var_party_label.set("Tercero (nombre/rut):")
        # Cambiar combo de estado segun informe
        if key in ("sales_period", "all_sales_orders"):
            self.cmb_state["values"] = self.SALES_STATES
            try: self.cmb_state.current(0)
            except Exception: pass
            if key == "all_sales_orders":
                self.var_party_label.set("Cliente (nombre/rut):")
        elif key in ("receivables_docs", "customer_debt"):
            self.cmb_state["values"] = self.RECEIVABLE_STATES
            try: self.cmb_state.current(0)
            except Exception: pass
            if key == "customer_debt":
                self.var_party_label.set("Cliente a consultar:")
        elif key in ("purchase_orders", "purchase_by_supplier_product", "all_purchase_orders"):
            self.cmb_state["values"] = self.PURCH_STATES
            try: self.cmb_state.current(0)
            except Exception: pass
            self.var_party_label.set("Proveedor (nombre/rut):")
        elif key in ("payables_docs", "supplier_debts"):
            self.cmb_state["values"] = self.PAYABLE_STATES
            try: self.cmb_state.current(0)
            except Exception: pass
        else:
            self.cmb_state["values"] = ["(Todos)"]
            try: self.cmb_state.current(0)
            except Exception: pass

        self._clear_current_report_data()

    def _clear_report_inputs(self) -> None:
        for var in (
            self.var_date_from,
            self.var_date_to,
            self.var_party,
            self.var_product,
            self.var_total_min,
            self.var_total_max,
        ):
            try:
                var.set("")
            except Exception:
                pass

    def _clear_current_report_data(self) -> None:
        self._current_cols, self._current_rows = [], []
        try:
            self.table.set_data([], [])
        except Exception:
            pass

    # ------------------------- Ejecución ------------------------- #
    def _run_report(self) -> None:
        try:
            idx = self.cmb_report.current() or 0
            key, _name = self.REPORTS[idx]
            if key == "stock_real":
                self._run_stock_report()
            elif key in ("sales_period", "receivables_docs"):
                self._run_sales_report(key)
            elif key == "all_sales_orders":
                self._run_all_sales_orders_report()
            elif key == "customer_debt":
                self._run_customer_debt_report()
            elif key == "all_purchase_orders":
                self._run_all_purchase_orders_report()
            elif key in ("purchase_orders", "purchase_by_supplier_product", "payables_docs"):
                self._run_purchases_report(key)
            elif key == "supplier_debts":
                self._run_supplier_debts_report()
            elif key == "price_list":
                self._run_price_list_report()
            elif key == "investment_by_product":
                self._run_investment_report()
            else:
                self._current_cols, self._current_rows = [], []
            # Cargar en tabla
            self.table.set_data(self._current_cols, self._current_rows)
            if key == "supplier_debts":
                try:
                    state_idx = self._current_cols.index("Estado")
                    self.table.set_row_backgrounds([
                        "#ffdddd" if len(row) > state_idx and str(row[state_idx]) == "Ingreso parcial" else None
                        for row in self._current_rows
                    ])
                except Exception:
                    pass
        except Exception as e:
            messagebox.showerror("Informes", f"No se pudo ejecutar el informe:\n{e}")

    # ---------------------- Stock real ---------------------- #
    def _run_stock_report(self) -> None:
        flt = InventoryFilter(report_type="completo")
        products = self.svc_inventory.fetch(flt)
        cols = ["ID", "Producto", "SKU", "Unidad", "Stock", "P. compra", "P. venta"]
        rows: List[List] = []
        for p in products:
            rows.append([
                p.id,
                p.nombre or "",
                p.sku or "",
                p.unidad_medida or "",
                int(p.stock_actual or 0),
                _fmt_money(p.precio_compra or 0),
                _fmt_money(p.precio_venta or 0),
            ])
        self._current_cols, self._current_rows = cols, rows

    def _run_price_list_report(self) -> None:
        cols = ["Producto", "Código", "Precio unitario"]
        q = self.session.query(Product)
        product_filter = (self.var_product.get() or "").strip()
        if product_filter:
            like = f"%{product_filter}%"
            q = q.filter((Product.nombre.ilike(like)) | (Product.sku.ilike(like)))

        rows: List[List] = []
        for p in q.order_by(Product.nombre.asc()).all():
            rows.append([
                p.nombre or "",
                p.sku or "",
                _fmt_money(p.precio_venta or 0),
            ])
        self._current_cols, self._current_rows = cols, rows

    # ------------------------ Ventas ------------------------ #
    def _run_sales_report(self, key: str) -> None:
        d_from, d_to = self._get_date_filters()
        state = (self.cmb_state.get() or "").strip()
        party_like = (self.var_party.get() or "").strip()
        def _num(s: str):
            try:
                return float((s or "").replace(".", "").replace(",", ".")) if s else None
            except Exception:
                return None
        tmin = _num(self.var_total_min.get()); tmax = _num(self.var_total_max.get())

        q = self.session.query(Sale, Customer).join(Customer, Customer.id == Sale.id_cliente)
        if d_from: q = q.filter(Sale.fecha_venta >= d_from)
        if d_to:   q = q.filter(Sale.fecha_venta <= d_to)

        if key == "receivables_docs":
            if not state or state == "(Todos)":
                q = q.filter(~Sale.estado.in_(["Pagado", "Pagada", "Confirmada"]))
            elif state == "Pagado":
                q = q.filter(Sale.estado.in_(["Pagado", "Pagada", "Confirmada"]))
            elif state == "Pendiente":
                q = q.filter(~Sale.estado.in_(["Pagado", "Pagada", "Confirmada"]))
            else:
                q = q.filter(Sale.estado == state)
        else:
            if state and state != "(Todos)":
                if state == "Pagado":
                    q = q.filter(Sale.estado.in_(["Pagado", "Pagada", "Confirmada"]))
                elif state == "Pendiente":
                    q = q.filter(~Sale.estado.in_(["Pagado", "Pagada", "Confirmada"]))
                else:
                    q = q.filter(Sale.estado == state)

        if party_like:
            like = f"%{party_like}%"
            q = q.filter((Customer.razon_social.ilike(like)) | (Customer.rut.ilike(like)))
        if tmin is not None: q = q.filter(Sale.total_venta >= tmin)
        if tmax is not None: q = q.filter(Sale.total_venta <= tmax)

        cols = ["ID", "Fecha", "Cliente", "Estado", "Total"]
        rows: List[List] = []
        for s, c in q.order_by(Sale.id.desc()):
            rows.append([
                s.id,
                s.fecha_venta.strftime("%Y-%m-%d %H:%M"),
                getattr(c, "razon_social", "") or "-",
                "Pagado" if str(s.estado or "").strip().lower() in ("pagado", "pagada", "confirmada") else "Pendiente",
                _fmt_money(s.total_venta or 0),
            ])
        self._current_cols, self._current_rows = cols, rows

    def _run_customer_debt_report(self) -> None:
        d_from, d_to = self._get_date_filters()
        state = (self.cmb_state.get() or "").strip()
        party_like = (self.var_party.get() or "").strip()

        def _num(s: str):
            try:
                return float((s or "").replace(".", "").replace(",", ".")) if s else None
            except Exception:
                return None

        tmin = _num(self.var_total_min.get()); tmax = _num(self.var_total_max.get())

        q = self.session.query(Sale, Customer).join(Customer, Customer.id == Sale.id_cliente)
        if d_from: q = q.filter(Sale.fecha_venta >= d_from)
        if d_to:   q = q.filter(Sale.fecha_venta <= d_to)

        if not state or state == "(Todos)":
            q = q.filter(~Sale.estado.in_(["Pagado", "Pagada", "Confirmada"]))
        elif state == "Pagado":
            q = q.filter(Sale.estado.in_(["Pagado", "Pagada", "Confirmada"]))
        elif state == "Pendiente":
            q = q.filter(~Sale.estado.in_(["Pagado", "Pagada", "Confirmada"]))
        else:
            q = q.filter(Sale.estado == state)

        if party_like:
            like = f"%{party_like}%"
            q = q.filter((Customer.razon_social.ilike(like)) | (Customer.rut.ilike(like)))
        if tmin is not None: q = q.filter(Sale.total_venta >= tmin)
        if tmax is not None: q = q.filter(Sale.total_venta <= tmax)

        cols = ["ID", "Factura", "Fecha", "Cliente", "Estado", "Estado Excel", "Neto", "IVA", "Total", "Fecha pagado", "Nota"]
        rows: List[List] = []
        total_deuda = 0.0
        for s, c in q.order_by(Customer.razon_social.asc(), Sale.fecha_venta.asc(), Sale.id.asc()):
            total = float(s.total_venta or 0)
            total_deuda += total
            fecha_pagado = getattr(s, "fecha_pagado", None)
            rows.append([
                s.id,
                getattr(s, "numero_documento", "") or "",
                s.fecha_venta.strftime("%Y-%m-%d %H:%M"),
                getattr(c, "razon_social", "") or "-",
                "Pagado" if str(s.estado or "").strip().lower() in ("pagado", "pagada", "confirmada") else "Pendiente",
                getattr(s, "estado_externo", "") or "",
                _fmt_money(getattr(s, "monto_neto", 0) or 0),
                _fmt_money(getattr(s, "monto_iva", 0) or 0),
                _fmt_money(total),
                fecha_pagado.strftime("%Y-%m-%d") if fecha_pagado else "",
                getattr(s, "nota", "") or "",
            ])
        if rows:
            rows.append(["", "", "", "TOTAL ADEUDADO", "", "", "", "", _fmt_money(total_deuda), "", ""])
        self._current_cols, self._current_rows = cols, rows

    def _run_all_sales_orders_report(self) -> None:
        d_from, d_to = self._get_date_filters()
        state = (self.cmb_state.get() or "").strip()
        party_like = (self.var_party.get() or "").strip()
        prod_like = (self.var_product.get() or "").strip()
        tmin = _parse_number(self.var_total_min.get())
        tmax = _parse_number(self.var_total_max.get())

        q = self.session.query(Sale, Customer).join(Customer, Customer.id == Sale.id_cliente)
        if d_from:
            q = q.filter(Sale.fecha_venta >= d_from)
        if d_to:
            q = q.filter(Sale.fecha_venta <= d_to)
        if state and state != "(Todos)":
            if state == "Pagado":
                q = q.filter(Sale.estado.in_(["Pagado", "Pagada", "Confirmada"]))
            elif state == "Pendiente":
                q = q.filter(~Sale.estado.in_(["Pagado", "Pagada", "Confirmada"]))
            else:
                q = q.filter(Sale.estado == state)
        if party_like:
            like = f"%{party_like}%"
            q = q.filter((Customer.razon_social.ilike(like)) | (Customer.rut.ilike(like)))
        if tmin is not None:
            q = q.filter(Sale.total_venta >= tmin)
        if tmax is not None:
            q = q.filter(Sale.total_venta <= tmax)

        sales = q.order_by(Sale.fecha_venta.desc(), Sale.id.desc()).all()
        sale_ids = [int(s.id) for s, _c in sales]
        product_rows: dict[int, list[tuple[SaleDetail, Product]]] = {}
        service_rows: dict[int, list[SaleServiceDetail]] = {}
        if sale_ids:
            detail_q = (
                self.session.query(SaleDetail, Product)
                .join(Product, Product.id == SaleDetail.id_producto)
                .filter(SaleDetail.id_venta.in_(sale_ids))
            )
            if prod_like:
                likep = f"%{prod_like}%"
                detail_q = detail_q.filter((Product.nombre.ilike(likep)) | (Product.sku.ilike(likep)))
            for det, prod in detail_q.order_by(SaleDetail.id_venta.asc(), Product.nombre.asc()).all():
                product_rows.setdefault(int(det.id_venta), []).append((det, prod))

            service_q = self.session.query(SaleServiceDetail).filter(SaleServiceDetail.id_venta.in_(sale_ids))
            if prod_like:
                service_q = service_q.filter(SaleServiceDetail.descripcion.ilike(f"%{prod_like}%"))
            for det in service_q.order_by(SaleServiceDetail.id_venta.asc(), SaleServiceDetail.descripcion.asc()).all():
                service_rows.setdefault(int(det.id_venta), []).append(det)

        cols = ["OV", "Fecha", "Cliente", "RUT", "Estado", "Doc", "Item", "SKU", "Cant.", "Precio", "Subtotal", "Total OV", "Nota"]
        rows: List[List] = []
        total_general = 0.0
        for sale, customer in sales:
            sid = int(sale.id)
            detail_items = product_rows.get(sid, [])
            service_items = service_rows.get(sid, [])
            if prod_like and not detail_items and not service_items:
                continue
            total_general += float(sale.total_venta or 0)
            base = [
                f"OV-{sid}",
                _fmt_date(sale.fecha_venta, True),
                getattr(customer, "razon_social", "") or "-",
                getattr(customer, "rut", "") or "",
                "Pagado" if str(sale.estado or "").strip().lower() in ("pagado", "pagada", "confirmada") else "Pendiente",
                getattr(sale, "numero_documento", "") or "",
            ]
            tail = [_fmt_money(sale.total_venta or 0), getattr(sale, "nota", "") or ""]
            wrote_row = False
            for det, prod in detail_items:
                rows.append(base + [
                    prod.nombre or "",
                    prod.sku or "",
                    float(det.cantidad or 0),
                    _fmt_money(det.precio_unitario or 0),
                    _fmt_money(det.subtotal or 0),
                ] + tail)
                wrote_row = True
            for det in service_items:
                rows.append(base + [
                    det.descripcion or "Servicio",
                    "SERV",
                    float(det.cantidad or 0),
                    _fmt_money(det.precio_unitario or 0),
                    _fmt_money(det.subtotal or 0),
                ] + tail)
                wrote_row = True
            if not wrote_row and not prod_like:
                rows.append(base + ["Sin detalle", "", "", "", "", *tail])
        if rows:
            rows.append(["", "", "", "", "", "", "", "", "", "", "TOTAL", _fmt_money(total_general), ""])
        self._current_cols, self._current_rows = cols, rows

    # ---------------------- Compras ---------------------- #
    def _purchase_paid_map(self, purchase_ids: list[int]) -> dict[int, float]:
        if not purchase_ids:
            return {}
        rows = (
            self.session.query(PurchasePayment.id_compra, func.coalesce(func.sum(PurchasePayment.monto), 0))
            .filter(PurchasePayment.id_compra.in_(purchase_ids))
            .group_by(PurchasePayment.id_compra)
            .all()
        )
        return {int(pid): float(total or 0) for pid, total in rows}

    def _run_all_purchase_orders_report(self) -> None:
        d_from, d_to = self._get_date_filters()
        state = (self.cmb_state.get() or "").strip()
        party_like = (self.var_party.get() or "").strip()
        prod_like = (self.var_product.get() or "").strip()
        tmin = _parse_number(self.var_total_min.get())
        tmax = _parse_number(self.var_total_max.get())

        q = (
            self.session.query(Purchase, PurchaseDetail, Product, Supplier)
            .join(PurchaseDetail, PurchaseDetail.id_compra == Purchase.id)
            .join(Product, Product.id == PurchaseDetail.id_producto)
            .join(Supplier, Supplier.id == Purchase.id_proveedor)
        )
        if d_from:
            q = q.filter(Purchase.fecha_compra >= d_from)
        if d_to:
            q = q.filter(Purchase.fecha_compra <= d_to)
        if state and state != "(Todos)":
            q = q.filter(Purchase.estado == state)
        if party_like:
            like = f"%{party_like}%"
            q = q.filter((Supplier.razon_social.ilike(like)) | (Supplier.rut.ilike(like)))
        if prod_like:
            likep = f"%{prod_like}%"
            q = q.filter((Product.nombre.ilike(likep)) | (Product.sku.ilike(likep)))
        if tmin is not None:
            q = q.filter(Purchase.total_compra >= tmin)
        if tmax is not None:
            q = q.filter(Purchase.total_compra <= tmax)

        cols = [
            "OC", "Fecha", "Proveedor", "RUT", "Estado", "Doc", "Venc.",
            "Producto", "SKU", "Cant.", "Recibida", "Precio", "Subtotal", "Total OC", "Politica stock",
        ]
        rows: List[List] = []
        total_general = 0.0
        seen_purchase_ids: set[int] = set()
        for pur, det, prod, sup in q.order_by(Purchase.fecha_compra.desc(), Purchase.id.desc(), Product.nombre.asc()).all():
            pid = int(pur.id)
            if pid not in seen_purchase_ids:
                total_general += float(pur.total_compra or 0)
                seen_purchase_ids.add(pid)
            rows.append([
                f"OC-{pid}",
                _fmt_date(pur.fecha_compra, True),
                getattr(sup, "razon_social", "") or "-",
                getattr(sup, "rut", "") or "",
                pur.estado or "",
                pur.numero_documento or "",
                _fmt_date(pur.fecha_vencimiento),
                prod.nombre or "",
                prod.sku or "",
                float(det.cantidad or 0),
                float(getattr(det, "received_qty", 0) or 0),
                _fmt_money(det.precio_unitario or 0),
                _fmt_money(det.subtotal or 0),
                _fmt_money(pur.total_compra or 0),
                getattr(pur, "stock_policy", "") or "",
            ])
        if rows:
            rows.append(["", "", "", "", "", "", "", "", "", "", "", "", "TOTAL", _fmt_money(total_general), ""])
        self._current_cols, self._current_rows = cols, rows

    def _run_supplier_debts_report(self) -> None:
        d_from, d_to = self._get_date_filters()
        state = (self.cmb_state.get() or "").strip()
        party_like = (self.var_party.get() or "").strip()

        q = self.session.query(Purchase, Supplier).join(Supplier, Supplier.id == Purchase.id_proveedor)
        if d_from:
            q = q.filter(Purchase.fecha_compra >= d_from)
        if d_to:
            q = q.filter(Purchase.fecha_compra <= d_to)
        if state and state != "(Todos)":
            q = q.filter(Purchase.estado == state)
        else:
            q = q.filter(Purchase.estado.in_(["Por pagar", "Ingreso parcial", "Pendiente", "Incompleta"]))
        if party_like:
            like = f"%{party_like}%"
            q = q.filter((Supplier.razon_social.ilike(like)) | (Supplier.rut.ilike(like)))

        purchases = q.order_by(Supplier.razon_social.asc(), Purchase.fecha_compra.asc(), Purchase.id.asc()).all()
        paid_by_purchase = self._purchase_paid_map([int(p.id) for p, _s in purchases])
        cols = ["Proveedor", "Compra", "Doc", "Fecha", "Estado", "Total", "Pagado", "Deuda", "Vencimiento"]
        rows: List[List] = []
        total_by_supplier: dict[int, float] = {}
        supplier_names: dict[int, str] = {}
        grand_total = 0.0

        for purchase, supplier in purchases:
            total = float(purchase.total_compra or 0)
            paid = float(paid_by_purchase.get(int(purchase.id), 0))
            debt = max(total - paid, 0.0)
            if debt <= 0 and str(purchase.estado or "") != "Ingreso parcial":
                continue
            supplier_id = int(supplier.id)
            supplier_name = getattr(supplier, "razon_social", "") or "-"
            supplier_names[supplier_id] = supplier_name
            total_by_supplier[supplier_id] = total_by_supplier.get(supplier_id, 0.0) + debt
            grand_total += debt
            rows.append([
                supplier_name,
                f"OC-{purchase.id}",
                purchase.numero_documento or "",
                purchase.fecha_compra.strftime("%Y-%m-%d") if purchase.fecha_compra else "",
                purchase.estado,
                _fmt_money(total),
                _fmt_money(paid),
                _fmt_money(debt),
                purchase.fecha_vencimiento.strftime("%Y-%m-%d") if purchase.fecha_vencimiento else "",
            ])

        if rows:
            rows.append(["", "", "", "", "", "", "", "", ""])
            for supplier_id, total_debt in sorted(total_by_supplier.items(), key=lambda item: supplier_names.get(item[0], "")):
                rows.append([supplier_names.get(supplier_id, ""), "TOTAL PROVEEDOR", "", "", "", "", "", _fmt_money(total_debt), ""])
            rows.append(["", "TOTAL GENERAL", "", "", "", "", "", _fmt_money(grand_total), ""])
        self._current_cols, self._current_rows = cols, rows

    def _run_purchases_report(self, key: str) -> None:
        d_from, d_to = self._get_date_filters()
        state = (self.cmb_state.get() or "").strip()
        party_like = (self.var_party.get() or "").strip()
        prod_like = (self.var_product.get() or "").strip()
        def _num(s: str):
            try:
                return float((s or "").replace(".", "").replace(",", ".")) if s else None
            except Exception:
                return None
        tmin = _num(self.var_total_min.get()); tmax = _num(self.var_total_max.get())

        def _fmt_date(val: Optional[dt.datetime], with_time: bool = False) -> str:
            if not val:
                return ""
            return val.strftime("%Y-%m-%d %H:%M") if with_time else val.strftime("%Y-%m-%d")

        if key == "purchase_by_supplier_product":
            q = (self.session.query(Purchase, PurchaseDetail, Product, Supplier)
                 .join(PurchaseDetail, PurchaseDetail.id_compra == Purchase.id)
                 .join(Product, Product.id == PurchaseDetail.id_producto)
                 .join(Supplier, Supplier.id == Purchase.id_proveedor))
            if d_from: q = q.filter(Purchase.fecha_compra >= d_from)
            if d_to:   q = q.filter(Purchase.fecha_compra <= d_to)
            if state and state != "(Todos)":
                q = q.filter(Purchase.estado == state)
            if party_like:
                like = f"%{party_like}%"
                q = q.filter((Supplier.razon_social.ilike(like)) | (Supplier.rut.ilike(like)))
            if prod_like:
                likep = f"%{prod_like}%"
                q = q.filter((Product.nombre.ilike(likep)) | (Product.sku.ilike(likep)))
            if tmin is not None: q = q.filter(Purchase.total_compra >= tmin)
            if tmax is not None: q = q.filter(Purchase.total_compra <= tmax)
            cols = ["Proveedor", "Compra ID", "Fecha", "Producto", "SKU", "Cant.", "Precio", "Subtotal", "Estado"]
            rows: List[List] = []
            for pur, det, prod, sup in q.order_by(Purchase.id.desc()):
                rows.append([
                    getattr(sup, "razon_social", "") or "-",
                    pur.id,
                    _fmt_date(pur.fecha_compra, True),
                    prod.nombre or "",
                    prod.sku or "",
                    float(det.cantidad or 0),
                    _fmt_money(det.precio_unitario or 0),
                    _fmt_money(det.subtotal or 0),
                    pur.estado,
                ])
            self._current_cols, self._current_rows = cols, rows
            return

        q = self.session.query(Purchase, Supplier).join(Supplier, Supplier.id == Purchase.id_proveedor)
        if d_from: q = q.filter(Purchase.fecha_compra >= d_from)
        if d_to:   q = q.filter(Purchase.fecha_compra <= d_to)

        if key == "payables_docs":
            if not state or state == "(Todos)":
                q = q.filter(Purchase.estado.in_(self.PAYABLE_STATES[1:]))
            else:
                q = q.filter(Purchase.estado == state)
        else:
            if state and state != "(Todos)":
                q = q.filter(Purchase.estado == state)

        if party_like:
            like = f"%{party_like}%"
            q = q.filter((Supplier.razon_social.ilike(like)) | (Supplier.rut.ilike(like)))
        if tmin is not None: q = q.filter(Purchase.total_compra >= tmin)
        if tmax is not None: q = q.filter(Purchase.total_compra <= tmax)

        if key == "payables_docs":
            purchases = q.order_by(Purchase.id.desc()).all()
            purchase_ids = [p.id for p, _s in purchases]
            docs_by_purchase: dict[int, list[str]] = {}
            if purchase_ids:
                for rec in (self.session.query(Reception)
                            .filter(Reception.id_compra.in_(purchase_ids))
                            .order_by(Reception.fecha.asc())):
                    label = f"{rec.tipo_doc or 'Doc'} {rec.numero_documento or ''}".strip()
                    docs_by_purchase.setdefault(rec.id_compra, []).append(label)

            cols = ["Compra ID", "Fecha", "Proveedor", "Estado", "Total", "Docs"]
            rows: List[List] = []
            for p, s in purchases:
                docs = docs_by_purchase.get(p.id) or []
                rows.append([
                    p.id,
                    _fmt_date(p.fecha_compra, True),
                    getattr(s, "razon_social", "") or "-",
                    p.estado,
                    _fmt_money(p.total_compra or 0),
                    " | ".join(docs) if docs else "Sin doc",
                ])
            self._current_cols, self._current_rows = cols, rows
            return

        cols = ["ID", "Fecha", "Proveedor", "Estado", "Total", "Doc", "F. doc", "Venc."]
        rows: List[List] = []
        for p, s in q.order_by(Purchase.id.desc()):
            rows.append([
                p.id,
                _fmt_date(p.fecha_compra, True),
                getattr(s, "razon_social", "") or "-",
                p.estado,
                _fmt_money(p.total_compra or 0),
                p.numero_documento or "",
                _fmt_date(p.fecha_documento),
                _fmt_date(p.fecha_vencimiento),
            ])
        self._current_cols, self._current_rows = cols, rows

    # -------------------- Inversion por producto -------------------- #
    def _run_investment_report(self) -> None:
        cols = ["ID", "Producto", "SKU", "Stock", "P. compra", "P. compra + IVA", "Inversion"]
        data: List[tuple[Decimal, List]] = []
        total = Decimal("0")
        for p in self.session.query(Product).order_by(Product.nombre.asc()).all():
            stock = int(p.stock_actual or 0)
            precio = Decimal(p.precio_compra or 0)
            precio_iva = precio * Decimal("1.19")
            inversion = precio_iva * Decimal(stock)
            total += inversion
            row = [
                p.id,
                p.nombre or "",
                p.sku or "",
                stock,
                _fmt_money(precio or 0),
                _fmt_money(precio_iva or 0),
                _fmt_money(inversion or 0),
            ]
            data.append((inversion, row))
        data.sort(key=lambda item: item[0], reverse=True)
        rows = [row for _inv, row in data]
        rows.append(["", "TOTAL", "", "", "", "", _fmt_money(total or 0)])
        self._current_cols, self._current_rows = cols, rows

    # -------------------- Exportar / Imprimir -------------------- #
    def _on_export(self) -> None:
        try:
            if not self._current_cols:
                self._run_report()
                self.table.set_data(self._current_cols, self._current_rows)
            if not self._current_cols:
                messagebox.showwarning("Exportar PDF", "No hay datos para exportar.")
                return
            path = self._export_current_pdf()
            webbrowser.open(str(path))
            messagebox.showinfo("OK", f"PDF generado:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar:\n{e}")

    def _on_export_xlsx(self) -> None:
        try:
            if not self._current_cols:
                self._run_report()
                self.table.set_data(self._current_cols, self._current_rows)
            if not self._current_cols:
                messagebox.showwarning("Exportar Excel", "No hay datos para exportar.")
                return
            path = self._export_current_xlsx()
            webbrowser.open(str(path))
            messagebox.showinfo("OK", f"Excel generado:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar Excel:\n{e}")

    def _export_current_pdf(self) -> Path:
        idx = self.cmb_report.current() or 0
        key, report_name = self.REPORTS[idx]
        timestamp = dt.datetime.now()
        out = _downloads_dir() / f"informe_{key}_{timestamp:%Y%m%d-%H%M%S}.pdf"
        company = _read_company_cfg()

        pages = self._render_pdf_pages(company, report_name, timestamp)
        if not pages:
            raise RuntimeError("No se pudo renderizar el PDF.")
        pages[0].save(str(out), "PDF", save_all=True, append_images=pages[1:], resolution=150)
        return out

    def _export_current_xlsx(self) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        idx = self.cmb_report.current() or 0
        key, report_name = self.REPORTS[idx]
        timestamp = dt.datetime.now()
        out = _downloads_dir() / f"informe_{key}_{timestamp:%Y%m%d-%H%M%S}.xlsx"
        company = _read_company_cfg()

        wb = Workbook()
        ws = wb.active
        ws.title = "Informe"
        max_col = max(1, len(self._current_cols))

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.cell(1, 1, report_name).font = Font(bold=True, size=15, color="0D2F53")
        ws.cell(1, 1).alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        ws.cell(2, 1, company.get("name") or "Inventario App").font = Font(bold=True, color="40566B")
        ws.cell(2, 1).alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
        ws.cell(3, 1, f"Generado: {timestamp:%d/%m/%Y %H:%M} | {self._pdf_filters_text()}")
        ws.cell(3, 1).alignment = Alignment(horizontal="center", wrap_text=True)

        header_row = 5
        head_fill = PatternFill("solid", fgColor="0D2F53")
        head_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="B8C6D5")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, title in enumerate(self._current_cols, start=1):
            cell = ws.cell(header_row, col_idx, title)
            cell.fill = head_fill
            cell.font = head_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row in enumerate(self._current_rows, start=header_row + 1):
            is_total = any(str(value).upper().startswith("TOTAL") for value in row)
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row_idx, col_idx, value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if is_total:
                    cell.fill = PatternFill("solid", fgColor="E9F2FB")
                    cell.font = Font(bold=True)

        ws.freeze_panes = "A6"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max(header_row, header_row + len(self._current_rows))}"
        for col_idx, title in enumerate(self._current_cols, start=1):
            values = [str(title)]
            values.extend(str(row[col_idx - 1]) for row in self._current_rows[:120] if col_idx - 1 < len(row))
            width = min(max(len(v) for v in values) + 2, 42)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 10)

        wb.save(out)
        return out

    def _render_pdf_pages(self, company: dict[str, str], report_name: str, timestamp: dt.datetime) -> list[Image.Image]:
        width, height = 1754, 1240
        margin = 70
        header_h = 178
        table_top = 250
        footer_h = 50
        fonts = self._pdf_fonts()
        col_widths = self._pdf_column_widths(width - 2 * margin, fonts["cell"])
        pages: list[Image.Image] = []

        def new_page() -> tuple[Image.Image, ImageDraw.ImageDraw, int]:
            image = Image.new("RGB", (width, height), "#FFFFFF")
            draw = ImageDraw.Draw(image)
            draw.rectangle([0, 0, width, height], fill="#F7FAFD")
            draw.rectangle([0, 0, width, 20], fill="#0D2F53")
            draw.rectangle([margin, 45, width - margin, 45 + header_h], fill="#FFFFFF", outline="#B8C6D5", width=2)
            self._draw_pdf_header(image, draw, company, report_name, timestamp, fonts, margin, width, 45)
            filters = self._pdf_filters_text()
            self._draw_wrapped(draw, filters, (margin, 45 + header_h + 18), width - 2 * margin, fonts["small"], "#40566B", 1)
            self._draw_pdf_record_count(draw, width - margin, 45 + header_h + 14, fonts)
            self._draw_pdf_table_header(draw, margin, table_top, col_widths, fonts)
            return image, draw, table_top + 42

        page, draw, y = new_page()
        pages.append(page)
        for row_idx, row in enumerate(self._current_rows):
            wrapped_cells = [
                self._wrap_text(str(row[idx] if idx < len(row) else ""), fonts["cell"], max(20, col_widths[idx] - 14))
                for idx in range(len(self._current_cols))
            ]
            row_h = max(32, max(len(lines) for lines in wrapped_cells) * 17 + 14)
            if y + row_h > height - footer_h - 20:
                page, draw, y = new_page()
                pages.append(page)
            fill = self._pdf_row_fill(row, row_idx)
            x = margin
            for idx, lines in enumerate(wrapped_cells):
                draw.rectangle([x, y, x + col_widths[idx], y + row_h], fill=fill, outline="#C4D0DC", width=1)
                ty = y + 7
                align = self._pdf_column_align(idx)
                for line in lines[: max(1, (row_h - 10) // 17)]:
                    if align == "right":
                        draw.text((x + col_widths[idx] - 8, ty), line, fill="#142B3F", font=fonts["cell"], anchor="ra")
                    elif align == "center":
                        draw.text((x + col_widths[idx] // 2, ty), line, fill="#142B3F", font=fonts["cell"], anchor="ma")
                    else:
                        draw.text((x + 7, ty), line, fill="#142B3F", font=fonts["cell"])
                    ty += 17
                x += col_widths[idx]
            y += row_h

        total_pages = len(pages)
        for idx, img in enumerate(pages, start=1):
            d = ImageDraw.Draw(img)
            d.text(
                (width - margin, height - 36),
                f"Página {idx} de {total_pages} - {timestamp:%d/%m/%Y %H:%M}",
                fill="#5D6C7A",
                font=fonts["small"],
                anchor="ra",
            )
            d.text((margin, height - 36), company.get("name") or "Inventario App", fill="#5D6C7A", font=fonts["small"])
        return pages

    def _pdf_fonts(self) -> dict[str, ImageFont.ImageFont]:
        def load(size: int, bold: bool = False):
            candidates = ["arialbd.ttf" if bold else "arial.ttf", "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf"]
            for name in candidates:
                try:
                    return ImageFont.truetype(name, size)
                except Exception:
                    continue
            return ImageFont.load_default()

        return {
            "title": load(34, True),
            "subtitle": load(18, False),
            "company": load(18, False),
            "company_bold": load(22, True),
            "small": load(15, False),
            "head": load(15, True),
            "cell": load(14, False),
        }

    def _draw_pdf_header(self, image, draw, company: dict[str, str], report_name: str, timestamp: dt.datetime, fonts, margin: int, width: int, top: int) -> None:
        logo_box = (margin + 18, top + 22, margin + 128, top + 132)
        draw.rectangle(logo_box, outline="#C4D0DC", width=1, fill="#F4F8FC")
        logo_path = (company.get("logo") or "").strip()
        if logo_path:
            path = Path(logo_path)
            if not path.is_absolute():
                path = Path.cwd() / path
            if path.exists():
                try:
                    logo = Image.open(path).convert("RGBA")
                    logo.thumbnail((106, 106))
                    px = logo_box[0] + (110 - logo.width) // 2
                    py = logo_box[1] + (110 - logo.height) // 2
                    image.paste(logo, (px, py), logo)
                except Exception:
                    draw.text((margin + 45, top + 68), "LOGO", fill="#687B8F", font=fonts["small"])
        else:
            draw.text((margin + 45, top + 68), "LOGO", fill="#687B8F", font=fonts["small"])

        x = margin + 150
        draw.text((x, top + 24), company.get("name") or "Inventario App", fill="#0A2B4F", font=fonts["company_bold"])
        company_lines = [
            f"RUT: {company.get('rut') or '-'}",
            company.get("address") or "",
            f"Tel: {company.get('phone') or '-'}",
            company.get("email") or "",
        ]
        y = top + 56
        for line in [item for item in company_lines if item]:
            draw.text((x, y), line, fill="#40566B", font=fonts["company"])
            y += 22

        draw.text((width - margin - 18, top + 30), report_name, fill="#0A2B4F", font=fonts["title"], anchor="ra")
        draw.text((width - margin - 18, top + 76), f"Generado el {timestamp:%d/%m/%Y %H:%M}", fill="#5D6C7A", font=fonts["subtitle"], anchor="ra")

    def _pdf_filters_text(self) -> str:
        parts = []
        if self.var_date_from.get().strip() or self.var_date_to.get().strip():
            parts.append(f"Periodo: {self.var_date_from.get().strip() or '-'} a {self.var_date_to.get().strip() or '-'}")
        state = (self.cmb_state.get() or "").strip()
        if state and state != "(Todos)":
            parts.append(f"Estado: {state}")
        if self.var_party.get().strip():
            parts.append(f"Tercero: {self.var_party.get().strip()}")
        if self.var_product.get().strip():
            parts.append(f"Producto: {self.var_product.get().strip()}")
        if self.var_total_min.get().strip() or self.var_total_max.get().strip():
            parts.append(f"Total: {self.var_total_min.get().strip() or '-'} a {self.var_total_max.get().strip() or '-'}")
        return " | ".join(parts) if parts else "Filtros: todos los registros disponibles."

    def _draw_pdf_record_count(self, draw, right_x: int, y: int, fonts) -> None:
        text = f"Registros: {len(self._current_rows)}"
        text_w = self._text_width(text, fonts["small"])
        x0 = right_x - text_w - 22
        draw.rectangle([x0, y - 4, right_x, y + 22], fill="#E9F2FB", outline="#B8C6D5", width=1)
        draw.text((right_x - 11, y), text, fill="#0D2F53", font=fonts["small"], anchor="ra")

    def _pdf_column_align(self, idx: int) -> str:
        if idx >= len(self._current_cols):
            return "left"
        name = str(self._current_cols[idx]).lower()
        if any(token in name for token in ("precio", "total", "subtotal", "saldo", "iva", "monto", "inversion")):
            return "right"
        if any(token in name for token in ("id", "codigo", "código", "sku", "fecha", "estado", "stock", "cant", "unidad")):
            return "center"
        return "left"

    def _pdf_row_fill(self, row: List, row_idx: int) -> str:
        if self._current_report_key == "supplier_debts":
            try:
                state_idx = self._current_cols.index("Estado")
                if len(row) > state_idx and str(row[state_idx]) == "Ingreso parcial":
                    return "#FFE1E1"
            except Exception:
                pass
        return "#FFFFFF" if row_idx % 2 == 0 else "#EEF4FA"

    def _pdf_column_widths(self, available_width: int, font: ImageFont.ImageFont) -> list[int]:
        cols = [str(c) for c in self._current_cols]
        if self._current_report_key == "price_list" and len(cols) == 3:
            widths = [int(available_width * 0.56), int(available_width * 0.22)]
            widths.append(available_width - sum(widths))
            return widths
        rows = [list(row) for row in self._current_rows]
        weights = []
        for idx, col in enumerate(cols):
            max_len = self._text_width(col, font)
            for row in rows[:80]:
                value = row[idx] if idx < len(row) else ""
                max_len = max(max_len, min(self._text_width(str(value), font), 360))
            weights.append(max(80, max_len))
        total_weight = sum(weights) or 1
        widths = [int(available_width * (w / total_weight)) for w in weights]
        diff = available_width - sum(widths)
        if widths:
            widths[-1] += diff
        return widths

    def _draw_pdf_table_header(self, draw, x: int, y: int, col_widths: list[int], fonts) -> None:
        cx = x
        for idx, title in enumerate(self._current_cols):
            draw.rectangle([cx, y, cx + col_widths[idx], y + 42], fill="#0D2F53", outline="#0D2F53")
            lines = self._wrap_text(str(title), fonts["head"], max(20, col_widths[idx] - 14))
            draw.text((cx + 7, y + 12), lines[0] if lines else "", fill="#FFFFFF", font=fonts["head"])
            cx += col_widths[idx]

    def _draw_wrapped(self, draw, text: str, pos: tuple[int, int], max_width: int, font, fill: str, max_lines: int = 2) -> None:
        y = pos[1]
        for line in self._wrap_text(text, font, max_width)[:max_lines]:
            draw.text((pos[0], y), line, fill=fill, font=font)
            y += 18

    @staticmethod
    def _text_width(text: str, font: ImageFont.ImageFont) -> int:
        try:
            box = font.getbbox(str(text))
            return int(box[2] - box[0])
        except Exception:
            return len(str(text)) * 8

    def _wrap_text(self, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
        text = (text or "").replace("\n", " ").strip()
        if not text:
            return [""]
        words = text.split()
        lines: list[str] = []
        line = ""
        for word in words:
            candidate = (line + " " + word).strip()
            if self._text_width(candidate, font) <= max_width:
                line = candidate
                continue
            if line:
                lines.append(line)
                line = word
            else:
                cut = word
                while cut and self._text_width(cut + "...", font) > max_width:
                    cut = cut[:-1]
                lines.append((cut + "...") if cut else word[:1])
                line = ""
        if line:
            lines.append(line)
        return lines or [""]

    def print_current(self) -> None:
        key = self._current_report_key or self.REPORTS[self.cmb_report.current() or 0][0]
        if key != "stock_real":
            messagebox.showinfo("Impresión", "Solo compatible con el informe de Stock real.")
            return
        # Usar impresora predeterminada si está configurada; de lo contrario preguntar
        printer_name = get_document_printer()
        if not printer_name:
            dlg = PrinterSelectDialog(self)
            self.wait_window(dlg)
            if not dlg.result:
                return
            printer_name = dlg.result
        try:
            flt = InventoryFilter(report_type="completo")
            path = print_inventory_report(self.session, flt, "Stock real", printer_name=printer_name)
            messagebox.showinfo("Impresión", f"Enviado a '{printer_name}'.\nArchivo: {path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo imprimir:\n{e}")


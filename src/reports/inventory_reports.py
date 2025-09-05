from __future__ import annotations
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Optional, Literal
import math
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from src.data.models import Product
from src.utils.helpers import get_inventory_limits
from .print_backend import print_xlsx as _print_xlsx

# 3 tipos de informe
ReportType = Literal["venta", "compra", "completo"]


# ---------------------------
# Filtro
# ---------------------------
@dataclass
class InventoryFilter:
    # Texto
    nombre_contains: Optional[str] = None
    sku_contains: Optional[str] = None
    unidad_equals: Optional[str] = None
    # IDs
    ids_in: Optional[Iterable[int]] = None
    # Stock
    stock_min: Optional[int] = None
    stock_max: Optional[int] = None
    solo_bajo_minimo: bool = False
    solo_sobre_maximo: bool = False
    # Precios (sobre el precio que corresponda al tipo de informe; en 'completo' se ignoran)
    precio_min: Optional[float] = None
    precio_max: Optional[float] = None
    # Orden
    order_by: str = "nombre"   # nombre | sku | stock | p_compra | p_venta
    order_asc: bool = True
    # Tipo de informe
    report_type: ReportType = "venta"


# ---------------------------
# Servicio
# ---------------------------
class InventoryReportService:
    """
    - 'venta'   : toda la info + SOLO columna P. Venta
    - 'compra'  : toda la info + SOLO columna P. Compra
    - 'completo': toda la info + P. Compra y P. Venta
    """

    def __init__(self, session: Session):
        self.session = session

    # ---- util ----
    @staticmethod
    def _is_num(x: Optional[float]) -> bool:
        try:
            return x is not None and not math.isnan(float(x))
        except Exception:
            return False

    def _query_products(self, flt: InventoryFilter) -> List[Product]:
        q = self.session.query(Product)

        # Texto
        if flt.ids_in:
            q = q.filter(Product.id.in_(list(flt.ids_in)))
        if flt.nombre_contains:
            q = q.filter(Product.nombre.ilike(f"%{flt.nombre_contains}%"))
        if flt.sku_contains:
            q = q.filter(Product.sku.ilike(f"%{flt.sku_contains}%"))
        if flt.unidad_equals:
            q = q.filter(Product.unidad_medida == flt.unidad_equals)

        # Stock
        crit_min, crit_max = get_inventory_limits()
        if flt.stock_min is not None:
            q = q.filter(Product.stock_actual >= flt.stock_min)
        if flt.stock_max is not None:
            q = q.filter(Product.stock_actual <= flt.stock_max)
        if flt.solo_bajo_minimo:
            q = q.filter(Product.stock_actual < crit_min)
        if flt.solo_sobre_maximo:
            q = q.filter(Product.stock_actual > crit_max)

        # Rango de precio (en 'completo' se ignora)
        if flt.report_type != "completo":
            if self._is_num(flt.precio_min):
                if flt.report_type == "venta":
                    q = q.filter(Product.precio_venta >= float(flt.precio_min))  # type: ignore[arg-type]
                else:
                    q = q.filter(Product.precio_compra >= float(flt.precio_min))  # type: ignore[arg-type]
            if self._is_num(flt.precio_max):
                if flt.report_type == "venta":
                    q = q.filter(Product.precio_venta <= float(flt.precio_max))  # type: ignore[arg-type]
                else:
                    q = q.filter(Product.precio_compra <= float(flt.precio_max))  # type: ignore[arg-type]

        # Orden (asegurar clave válida)
        colmap = {
            "nombre": Product.nombre,
            "sku": Product.sku,
            "stock": Product.stock_actual,
            "p_compra": Product.precio_compra,
            "p_venta": Product.precio_venta,
        }
        order_key = flt.order_by if flt.order_by in colmap else "nombre"
        col = colmap[order_key]
        q = q.order_by(col.asc() if flt.order_asc else col.desc())
        return q.all()

    def fetch(self, flt: InventoryFilter) -> List[Product]:
        return self._query_products(flt)

    def export_xlsx(self, rows: List[Product], flt: InventoryFilter, title: str) -> Path:
        """
        Estructura base:
            ID | Producto | SKU | Unidad | Stock | precio(s)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        titulo_tipo = {"venta": "VENTA", "compra": "COMPRA", "completo": "COMPLETO"}[flt.report_type]
        ws["A1"] = f"{title} ({titulo_tipo})"
        ws["A2"] = f"Generado: {now}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"].font = Font(italic=True, size=10)

        # Encabezados
        headers = ["ID", "Producto", "SKU", "Unidad", "Stock"]
        if flt.report_type == "venta":
            headers.append("P. Venta")
        elif flt.report_type == "compra":
            headers.append("P. Compra")
        else:  # completo
            headers += ["P. Compra", "P. Venta"]

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.append(headers)
        header_row = 3
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="DDDDDD")

        # estilos de filas
        crit_min, crit_max = get_inventory_limits()
        red = PatternFill("solid", fgColor="FFDDDD")
        yellow = PatternFill("solid", fgColor="FFF6CC")
        thin = Side(border_style="thin", color="999999")

        # Filas
        for p in rows:
            stock = int(p.stock_actual or 0)
            row = [p.id, p.nombre, p.sku, p.unidad_medida or "", stock]

            if flt.report_type == "venta":
                row.append(float(p.precio_venta or 0.0))
            elif flt.report_type == "compra":
                row.append(float(p.precio_compra or 0.0))
            else:
                row += [float(p.precio_compra or 0.0), float(p.precio_venta or 0.0)]

            ws.append(row)

            r = ws.max_row
            fill = red if stock < crit_min else (yellow if stock > crit_max else None)
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c)
                if fill:
                    cell.fill = fill
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if c in (1, 4, 5):  # ID, Unidad, Stock
                    cell.alignment = Alignment(horizontal="center")
                if headers[c - 1] in ("P. Compra", "P. Venta"):
                    cell.number_format = "#,##0.00"

        # Auto ancho
        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            max_len = max(len(str(cell.value or "")) for cell in ws[letter])
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 50)

        # Config de impresión
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = f"{header_row}:{header_row}"

        out = Path(tempfile.gettempdir()) / f"Inventario_{flt.report_type}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        wb.save(out)
        return out


# ---------------------------
# Wrappers
# ---------------------------
def generate_inventory_xlsx(session: Session, flt: InventoryFilter, title: str) -> Path:
    svc = InventoryReportService(session)
    rows = svc.fetch(flt)
    return svc.export_xlsx(rows, flt, title)

def print_inventory_report(session: Session, flt: InventoryFilter, title: str, printer_name: str | None = None) -> Path:
    xlsx = generate_inventory_xlsx(session, flt, title)
    _print_xlsx(xlsx, printer_name=printer_name)
    return xlsx

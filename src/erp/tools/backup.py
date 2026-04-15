from __future__ import annotations
from pathlib import Path
from typing import Iterable, List, Dict, Any, Optional
import sqlite3
import base64
import shutil
from datetime import datetime

from openpyxl import Workbook, load_workbook

from src.erp.core.database import get_connection, init_db, DEFAULT_DB_PATH
from src.data.database import get_session
from src.data.models import (
    Product, Customer, Supplier,
    Purchase, PurchaseDetail,
    Sale, SaleDetail, SaleServiceDetail,
    Reception,
    StockEntry, StockExit,
    Location,
)
from src.utils.helpers import get_downloads_dir, unique_path
from src.utils.image_store import PRODUCTS_DIR


def _rows(conn: sqlite3.Connection, sql: str) -> List[sqlite3.Row]:
    cur = conn.execute(sql)
    return list(cur.fetchall())


def export_erp_to_xlsx(out_path: Optional[Path] = None, *, auto_open: bool = True) -> Path:
    # Usar el mismo archivo que emplea la app (DEFAULT_DB_PATH)
    conn = get_connection(DEFAULT_DB_PATH)
    try:
        init_db(conn)
    except Exception:
        pass

    docs = _rows(conn, "SELECT * FROM documentos ORDER BY id ASC")
    dets = _rows(conn, "SELECT * FROM detalles ORDER BY id ASC")
    logs = _rows(conn, "SELECT * FROM log_auditoria ORDER BY id ASC")

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "meta"
    ws_meta["A1"] = "generated_at"
    ws_meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _write_sheet(name: str, rows: List[sqlite3.Row]):
        ws = wb.create_sheet(title=name)
        if not rows:
            return
        cols = rows[0].keys()
        ws.append(list(cols))
        for r in rows:
            ws.append([r[c] for c in cols])

    _write_sheet("documentos", docs)
    _write_sheet("detalles", dets)
    _write_sheet("log_auditoria", logs)

    # save
    if out_path is None:
        out_dir = get_downloads_dir()
        out_path = unique_path(out_dir, f"ERP_backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    if auto_open:
        try:
            import webbrowser
            webbrowser.open(str(out_path))
        except Exception:
            pass
    return out_path


def import_erp_from_xlsx(xlsx_path: Path) -> None:
    if not Path(xlsx_path).exists():
        raise FileNotFoundError(xlsx_path)
    # Importar sobre el mismo archivo ERP usado por la app
    conn = get_connection(DEFAULT_DB_PATH)
    init_db(conn)
    wb = load_workbook(filename=str(xlsx_path))

    def _read_sheet(name: str) -> List[Dict[str, Any]]:
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = list(ws.values)
        if not rows:
            return []
        headers = [str(h) for h in rows[0]]
        out: List[Dict[str, Any]] = []
        for r in rows[1:]:
            if r is None:
                continue
            rec = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
            out.append(rec)
        return out

    documentos = _read_sheet("documentos")
    detalles = _read_sheet("detalles")
    logs = _read_sheet("log_auditoria")

    cur = conn.cursor()
    cur.execute("PRAGMA foreign_keys = ON;")
    # limpiar en orden seguro
    cur.execute("DELETE FROM detalles;")
    cur.execute("DELETE FROM documentos;")
    cur.execute("DELETE FROM log_auditoria;")

    # insertar documentos
    if documentos:
        cols = list(documentos[0].keys())
        placeholders = ",".join([":" + c for c in cols])
        sql = f"INSERT INTO documentos ({','.join(cols)}) VALUES ({placeholders})"
        cur.executemany(sql, documentos)

    # insertar detalles (dependen de documentos)
    if detalles:
        cols = list(detalles[0].keys())
        placeholders = ",".join([":" + c for c in cols])
        sql = f"INSERT INTO detalles ({','.join(cols)}) VALUES ({placeholders})"
        cur.executemany(sql, detalles)

    # logs
    if logs:
        cols = list(logs[0].keys())
        placeholders = ",".join([":" + c for c in cols])
        sql = f"INSERT INTO log_auditoria ({','.join(cols)}) VALUES ({placeholders})"
        cur.executemany(sql, logs)

    conn.commit()


# =================== APP BACKUP (ORM) =================== #
def _to_val(v):
    try:
        from decimal import Decimal
        if isinstance(v, Decimal):
            return float(v)
        return v
    except Exception:
        return v


def _to_float(v) -> Optional[float]:
    """Coerce Excel cell value into float robustly.
    - Accepts Python numbers/Decimal
    - Parses strings with currency symbols, thousands separators and decimal comma
      Examples accepted: "1.234,56" → 1234.56, "1,234.56" → 1234.56, "$ 1.234" → 1234.0
    Returns None if cannot parse.
    """
    try:
        from decimal import Decimal
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, Decimal):
            return float(v)
        s = str(v).strip()
        if not s:
            return None
        import re
        # keep only digits, separators, minus
        s = re.sub(r"[^0-9,\.\-]", "", s)
        if not s:
            return None
        # Determine decimal separator strategy
        if "," in s and "." in s:
            # If last comma is after last dot -> comma is decimal sep
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s:
            # treat comma as decimal separator
            s = s.replace(",", ".")
        # else: dot is decimal separator or integer
        return float(s)
    except Exception:
        return None


def _to_bool(v, default: bool = False) -> bool:
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    text = str(v).strip().lower()
    if text in {"1", "true", "si", "sí", "yes", "y"}:
        return True
    if text in {"0", "false", "no", "n"}:
        return False
    return default


def _iter_product_media_rows(products: list[Product]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for p in products:
        try:
            product_media_dir = PRODUCTS_DIR / str(int(p.id))
        except Exception:
            continue
        if not product_media_dir.exists():
            continue

        image_path_value = str(getattr(p, "image_path", None) or "").strip()
        image_path_name = Path(image_path_value).name if image_path_value else ""

        for media_file in sorted(product_media_dir.rglob("*")):
            if not media_file.is_file():
                continue
            try:
                rel_path = media_file.relative_to(PRODUCTS_DIR).as_posix()
                payload = base64.b64encode(media_file.read_bytes()).decode("ascii")
                is_primary = "1" if (
                    image_path_value and (
                        str(media_file.resolve()) == image_path_value or
                        media_file.name == image_path_name
                    )
                ) else "0"
                rows.append([
                    int(p.id),
                    rel_path,
                    media_file.suffix.lower(),
                    is_primary,
                    payload,
                ])
            except Exception:
                continue
    return rows


def export_app_backup_to_xlsx(out_path: Optional[Path] = None, *, auto_open: bool = True) -> Path:
    """Exporta un backup APP completo para restauracion precisa."""
    sess = get_session()

    # 1) Crear libro y meta
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "meta"
    ws_meta["A1"] = "generated_at"
    ws_meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 2) Ubicaciones
    ws = wb.create_sheet("ubicaciones")
    loc_cols = ["id", "nombre", "descripcion"]
    ws.append(loc_cols)
    for l in sess.query(Location).order_by(Location.id.asc()).all():
        ws.append([l.id, l.nombre, getattr(l, 'descripcion', None) or ""])

    # 3) Productos
    ws = wb.create_sheet("productos")
    prod_cols = [
        "id","nombre","sku","precio_compra","precio_venta","stock_actual",
        "unidad_medida","familia","barcode","id_proveedor","id_ubicacion","image_path"
    ]
    ws.append(prod_cols)
    products = sess.query(Product).order_by(Product.id.asc()).all()
    for p in products:
        ws.append([
            p.id, p.nombre, p.sku,
            _to_val(getattr(p, 'precio_compra', 0)),
            _to_val(getattr(p, 'precio_venta', 0)),
            getattr(p, 'stock_actual', 0),
            getattr(p, 'unidad_medida', None) or "",
            getattr(p, 'familia', None) or "",
            getattr(p, 'barcode', None) or "",
            getattr(p, 'id_proveedor', None),
            getattr(p, 'id_ubicacion', None),
            getattr(p, 'image_path', None) or "",
        ])

    # 3b) Imágenes de productos
    ws = wb.create_sheet("imagenes_productos")
    image_cols = ["id_producto", "ruta_relativa", "extension", "es_principal", "archivo_base64"]
    ws.append(image_cols)
    for row in _iter_product_media_rows(products):
        ws.append(row)

    # 4) Clientes
    ws = wb.create_sheet("clientes")
    cust_cols = ["id","razon_social","rut","contacto","telefono","email","direccion"]
    ws.append(cust_cols)
    for c in sess.query(Customer).order_by(Customer.id.asc()).all():
        ws.append([
            c.id, c.razon_social, c.rut,
            getattr(c,'contacto',None) or "",
            getattr(c,'telefono',None) or "",
            getattr(c,'email',None) or "",
            getattr(c,'direccion',None) or "",
        ])

    # 5) Proveedores
    ws = wb.create_sheet("proveedores")
    sup_cols = ["id","razon_social","rut","contacto","telefono","email","direccion"]
    ws.append(sup_cols)
    for s in sess.query(Supplier).order_by(Supplier.id.asc()).all():
        ws.append([
            s.id, s.razon_social, s.rut,
            getattr(s,'contacto',None) or "",
            getattr(s,'telefono',None) or "",
            getattr(s,'email',None) or "",
            getattr(s,'direccion',None) or "",
        ])

    # 6) Órdenes (todas: compras, ventas y recepciones)
    ws = wb.create_sheet("ordenes")
    order_cols = [
        # genéricos
        "tipo", "id", "fecha", "estado", "total",
        # tercero
        "tercero_tipo", "tercero_id", "tercero_nombre",
        # documento asociado (si aplica)
        "doc_tipo", "doc_numero",
        # vínculos específicos
        "oc_id",    # para compras/recepciones
        "ov_id",    # para ventas
        # campos adicionales de compra
        "referencia", "moneda", "tasa_cambio", "unidad_negocio", "proporcionalidad", "stock_policy",
        "fecha_documento", "fecha_contable", "fecha_vencimiento", "atencion",
        "tipo_descuento", "descuento", "ajuste_iva", "ajuste_impuesto",
        "numero_documento", "mes_referencia", "monto_neto", "monto_iva", "fecha_pagado", "nota",
        "estado_externo", "origen"
    ]
    ws.append(order_cols)

    # Compras
    for pur, sup in (
        sess.query(Purchase, Supplier)
        .join(Supplier, Supplier.id == Purchase.id_proveedor)
        .order_by(Purchase.id.asc())
        .all()
    ):
        ws.append([
            "compra",
            pur.id,
            getattr(pur, 'fecha_compra', None),
            getattr(pur, 'estado', None) or "",
            _to_val(getattr(pur, 'total_compra', 0)),
            "proveedor",
            getattr(pur, 'id_proveedor', None),
            getattr(sup, 'razon_social', None) or "",
            "",
            getattr(pur, 'numero_documento', None) or "",
            getattr(pur, 'id', None),   # oc_id
            None,                       # ov_id vacío
            getattr(pur, 'referencia', None) or "",
            getattr(pur, 'moneda', None) or "",
            _to_val(getattr(pur, 'tasa_cambio', None)),
            getattr(pur, 'unidad_negocio', None) or "",
            getattr(pur, 'proporcionalidad', None) or "",
            getattr(pur, 'stock_policy', None) or "",
            getattr(pur, 'fecha_documento', None),
            getattr(pur, 'fecha_contable', None),
            getattr(pur, 'fecha_vencimiento', None),
            getattr(pur, 'atencion', None) or "",
            getattr(pur, 'tipo_descuento', None) or "",
            _to_val(getattr(pur, 'descuento', None)),
            _to_val(getattr(pur, 'ajuste_iva', None)),
            _to_val(getattr(pur, 'ajuste_impuesto', None)),
            "",
            "",
            None,
            None,
            None,
            "",
            "",
            "",
        ])

    # Ventas
    for sale, cust in (
        sess.query(Sale, Customer)
        .join(Customer, Customer.id == Sale.id_cliente)
        .order_by(Sale.id.asc())
        .all()
    ):
        ws.append([
            "venta",
            sale.id,
            getattr(sale, 'fecha_venta', None),
            getattr(sale, 'estado', None) or "",
            _to_val(getattr(sale, 'total_venta', 0)),
            "cliente",
            getattr(sale, 'id_cliente', None),
            getattr(cust, 'razon_social', None) or getattr(cust, 'rut', None) or "",
            "",
            "",
            None,               # oc_id
            getattr(sale, 'id', None),  # ov_id
            "",
            "",
            None,
            "",
            "",
            "",
            None,
            None,
            None,
            "",
            "",
            None,
            None,
            None,
            getattr(sale, 'numero_documento', None) or "",
            getattr(sale, 'mes_referencia', None) or "",
            _to_val(getattr(sale, 'monto_neto', None)),
            _to_val(getattr(sale, 'monto_iva', None)),
            getattr(sale, 'fecha_pagado', None),
            getattr(sale, 'nota', None) or "",
            getattr(sale, 'estado_externo', None) or "",
            getattr(sale, 'origen', None) or "",
        ])

    # Recepciones
    for rec, pur, sup in (
        sess.query(Reception, Purchase, Supplier)
        .join(Purchase, Purchase.id == Reception.id_compra)
        .join(Supplier, Supplier.id == Purchase.id_proveedor)
        .order_by(Reception.id.asc())
        .all()
    ):
        ws.append([
            "recepcion",
            rec.id,
            getattr(rec, 'fecha', None),
            getattr(pur, 'estado', None) or "",
            _to_val(getattr(pur, 'total_compra', 0)),
            "proveedor",
            getattr(pur, 'id_proveedor', None),
            getattr(sup, 'razon_social', None) or "",
            getattr(rec, 'tipo_doc', None) or "",
            getattr(rec, 'numero_documento', None) or "",
            getattr(pur, 'id', None),   # oc_id
            None,                       # ov_id
            getattr(pur, 'referencia', None) or "",
            getattr(pur, 'moneda', None) or "",
            _to_val(getattr(pur, 'tasa_cambio', None)),
            getattr(pur, 'unidad_negocio', None) or "",
            getattr(pur, 'proporcionalidad', None) or "",
            getattr(pur, 'stock_policy', None) or "",
            getattr(pur, 'fecha_documento', None),
            getattr(pur, 'fecha_contable', None),
            getattr(pur, 'fecha_vencimiento', None),
            getattr(pur, 'atencion', None) or "",
            getattr(pur, 'tipo_descuento', None) or "",
            _to_val(getattr(pur, 'descuento', None)),
            _to_val(getattr(pur, 'ajuste_iva', None)),
            _to_val(getattr(pur, 'ajuste_impuesto', None)),
            "",
            "",
            None,
            None,
            None,
            "",
            "",
            "",
        ])

    # 6b) Detalles de compras
    ws = wb.create_sheet("detalles_compra")
    ws.append(["id", "id_compra", "id_producto", "cantidad", "received_qty", "precio_unitario", "subtotal"])
    for det in sess.query(PurchaseDetail).order_by(PurchaseDetail.id.asc()).all():
        ws.append([
            det.id,
            det.id_compra,
            det.id_producto,
            det.cantidad,
            getattr(det, "received_qty", 0),
            _to_val(det.precio_unitario),
            _to_val(det.subtotal),
        ])

    # 6c) Detalles de ventas
    ws = wb.create_sheet("detalles_venta")
    ws.append(["id", "id_venta", "id_producto", "cantidad", "precio_unitario", "subtotal"])
    for det in sess.query(SaleDetail).order_by(SaleDetail.id.asc()).all():
        ws.append([
            det.id,
            det.id_venta,
            det.id_producto,
            det.cantidad,
            _to_val(det.precio_unitario),
            _to_val(det.subtotal),
        ])

    # 6d) Servicios manuales de ventas
    ws = wb.create_sheet("servicios_venta")
    ws.append(["id", "id_venta", "descripcion", "cantidad", "precio_unitario", "subtotal", "afecto_iva"])
    for det in sess.query(SaleServiceDetail).order_by(SaleServiceDetail.id.asc()).all():
        ws.append([
            det.id,
            det.id_venta,
            det.descripcion,
            det.cantidad,
            _to_val(det.precio_unitario),
            _to_val(det.subtotal),
            "1" if getattr(det, "afecto_iva", True) else "0",
        ])

    # 7) Inventario (snapshot de cantidades por producto)
    # Nota: el usuario solicitó que sea por cantidad (no movimientos E/S)
    ws = wb.create_sheet("inventario")
    inv_cols = ["id_producto", "sku", "nombre", "stock_actual", "id_ubicacion"]
    ws.append(inv_cols)
    for p in sess.query(Product).order_by(Product.id.asc()).all():
        ws.append([
            p.id,
            getattr(p, 'sku', None) or "",
            getattr(p, 'nombre', None) or "",
            getattr(p, 'stock_actual', 0) or 0,
            getattr(p, 'id_ubicacion', None),
        ])

    # 8) Movimientos de stock
    ws = wb.create_sheet("movimientos_entrada")
    ws.append([
        "id", "id_producto", "id_ubicacion", "id_recepcion", "cantidad", "motivo",
        "lote", "serie", "fecha_vencimiento", "fecha",
    ])
    for mov in sess.query(StockEntry).order_by(StockEntry.id.asc()).all():
        ws.append([
            mov.id,
            mov.id_producto,
            getattr(mov, "id_ubicacion", None),
            getattr(mov, "id_recepcion", None),
            mov.cantidad,
            getattr(mov, "motivo", None) or "",
            getattr(mov, "lote", None) or "",
            getattr(mov, "serie", None) or "",
            getattr(mov, "fecha_vencimiento", None),
            getattr(mov, "fecha", None),
        ])

    ws = wb.create_sheet("movimientos_salida")
    ws.append([
        "id", "id_producto", "id_ubicacion", "cantidad", "motivo", "lote", "serie", "fecha",
    ])
    for mov in sess.query(StockExit).order_by(StockExit.id.asc()).all():
        ws.append([
            mov.id,
            mov.id_producto,
            getattr(mov, "id_ubicacion", None),
            mov.cantidad,
            getattr(mov, "motivo", None) or "",
            getattr(mov, "lote", None) or "",
            getattr(mov, "serie", None) or "",
            getattr(mov, "fecha", None),
        ])

    # Guardar
    if out_path is None:
        out_dir = get_downloads_dir()
        out_path = unique_path(out_dir, f"APP_backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    if auto_open:
        try:
            import webbrowser
            webbrowser.open(str(out_path))
        except Exception:
            pass
    return out_path


def export_app_quick_load_template_to_xlsx(
    out_path: Optional[Path] = None,
    *,
    auto_open: bool = True,
) -> Path:
    """
    Genera una plantilla de cargado rápido basada en la estructura real de la APP.

    La plantilla:
    - exporta las hojas que la importación automática ya sabe leer,
    - agrega una hoja inicial con instrucciones de uso,
    - permite que el usuario copie/pegue columnas desde sus propias planillas
      y luego reimporte el archivo directamente.
    """
    out = export_app_backup_to_xlsx(out_path=out_path, auto_open=False)
    wb = load_workbook(filename=str(out))

    if "instrucciones" in wb.sheetnames:
        ws = wb["instrucciones"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("instrucciones", 0)

    instructions = [
        ["Plantilla de cargado rápido - Inventario App"],
        [""],
        ["Objetivo"],
        [
            "Usa este archivo para cargar datos base de la empresa desde planillas propias. "
            "Puedes copiar, mover o adaptar columnas dentro de las hojas editables y luego "
            "usar la opción de importación automática de plantilla."
        ],
        [""],
        ["Hojas editables recomendadas"],
        ["ubicaciones", "Bodegas, salas, sucursales o zonas internas."],
        ["proveedores", "Datos maestros de proveedores."],
        ["clientes", "Datos maestros de clientes."],
        ["productos", "Catálogo principal de productos."],
        ["inventario", "Stock final por producto si quieres cargar cantidades actuales."],
        ["ordenes", "Opcional. Para cabeceras históricas simplificadas de compras, ventas y recepciones."],
        [""],
        ["Reglas importantes"],
        ["1.", "No cambies los nombres de las hojas."],
        ["2.", "Mantén los encabezados de la primera fila."],
        ["3.", "Si no usarás una hoja, puedes dejarla vacía."],
        ["4.", "Si importas con reemplazo total, la app tomará esta plantilla como base de la empresa."],
        ["5.", "Facturion no se mezcla con esta plantilla; su base de datos sigue siendo independiente."],
        [""],
        ["Consejo de uso"],
        [
            "Si tu cliente ya tiene una plantilla propia, copia sus columnas hacia estas hojas "
            "respetando los encabezados. Luego importa este mismo archivo desde la app."
        ],
    ]
    for row in instructions:
        ws.append(row)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 120
    try:
        ws.freeze_panes = "A3"
    except Exception:
        pass

    for sheet_name in ("ubicaciones", "proveedores", "clientes", "productos", "ordenes", "inventario"):
        if sheet_name not in wb.sheetnames:
            continue
        sh = wb[sheet_name]
        try:
            sh.freeze_panes = "A2"
        except Exception:
            pass
        for column_cells in sh.columns:
            try:
                letter = column_cells[0].column_letter
                max_len = max(len(str(cell.value or "")) for cell in column_cells[:25])
                sh.column_dimensions[letter].width = min(max(max_len + 2, 12), 36)
            except Exception:
                pass

    wb.save(out)
    try:
        wb.close()
    except Exception:
        pass

    if auto_open:
        try:
            import webbrowser
            webbrowser.open(str(out))
        except Exception:
            pass
    return out


def import_app_quick_load_template_from_xlsx(xlsx_path: Path, *, reset: bool = True) -> None:
    """
    Importa una plantilla de cargado rápido.

    Internamente usa el mismo formato del backup APP para asegurar compatibilidad.
    """
    import_app_backup_from_xlsx(xlsx_path, reset=reset)


def import_app_backup_from_xlsx(xlsx_path: Path, *, reset: bool = True) -> None:
    """
    Importa un backup APP generado por export_app_backup_to_xlsx.

    Estrategia:
    - (Opcional) Limpia tablas dependientes y maestras respetando FKs.
    - Inserta maestros (ubicaciones, proveedores, clientes, productos).
    - Inserta cabeceras de órdenes, luego detalles y servicios.
    - Restaura imágenes y movimientos de stock si están presentes.
    - Aplica snapshot final de stock desde la hoja "inventario".

    También mantiene compatibilidad con backups antiguos que no traían todas las hojas.
    """
    p = Path(xlsx_path)
    if not p.exists():
        raise FileNotFoundError(p)

    wb = load_workbook(filename=str(p))

    def _read(name: str) -> list[dict]:
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = list(ws.values)
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        out: list[dict] = []
        for r in rows[1:]:
            if r is None:
                continue
            out.append({headers[i]: r[i] if i < len(r) else None for i in range(len(headers))})
        return out

    sess = get_session()
    try:
        if reset:
            # Orden seguro respetando FKs
            sess.query(StockEntry).delete()
            sess.query(StockExit).delete()
            sess.query(SaleServiceDetail).delete()
            sess.query(SaleDetail).delete()
            sess.query(PurchaseDetail).delete()
            try:
                sess.query(Reception).delete()
            except Exception:
                pass
            sess.query(Sale).delete()
            sess.query(Purchase).delete()
            sess.query(Product).delete()
            sess.query(Supplier).delete()
            sess.query(Customer).delete()
            # Eliminar ubicaciones al final, cuando no quedan productos referenciándolas
            try:
                sess.query(Location).delete()
            except Exception:
                pass
            sess.commit()
            try:
                if PRODUCTS_DIR.exists():
                    shutil.rmtree(PRODUCTS_DIR)
            except Exception:
                pass
        try:
            PRODUCTS_DIR.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

        # Ubicaciones
        for r in _read("ubicaciones"):
            try:
                loc = Location(
                    id=int(r.get("id")),
                    nombre=str(r.get("nombre") or f"Ubicacion {r.get('id')}") or f"Ubicacion {r.get('id')}",
                    descripcion=r.get("descripcion") or None,
                )
                sess.add(loc)
            except Exception:
                pass
        sess.flush()

        # Proveedores (garantizar unicidad por id y rut)
        seen_sup_ids: set[int] = set()
        seen_sup_ruts: set[str] = set()
        for r in _read("proveedores"):
            try:
                sid = int(r.get("id"))
                if sid in seen_sup_ids:
                    continue
                seen_sup_ids.add(sid)

                rs = str(r.get("razon_social") or "")
                rut_raw = str(r.get("rut") or "").strip()
                rut_val = rut_raw if rut_raw else f"FAKE-{sid}"
                if rut_val in seen_sup_ruts:
                    rut_val = f"{rut_val}-{sid}"
                seen_sup_ruts.add(rut_val)

                sup = Supplier(
                    id=sid,
                    razon_social=rs,
                    rut=rut_val,
                    contacto=r.get("contacto"), telefono=r.get("telefono"),
                    email=r.get("email"), direccion=r.get("direccion"),
                )
                sess.add(sup)
            except Exception:
                pass
        sess.flush()

        # Clientes
        for r in _read("clientes"):
            try:
                cust = Customer(
                    id=int(r.get("id")),
                    razon_social=str(r.get("razon_social") or r.get("nombre") or ""),
                    rut=str(r.get("rut") or f"FAKE-{r.get('id')}") or f"FAKE-{r.get('id')}",
                    contacto=r.get("contacto"), telefono=r.get("telefono"),
                    email=r.get("email"), direccion=r.get("direccion"),
                )
                sess.add(cust)
            except Exception:
                pass
        sess.flush()

        # Cache de ubicaciones existentes (ids)
        existing_locations = {int(l.id) for l in sess.query(Location.id).all()} if sess.bind else set()

        # Productos (asegurando unicidad de SKU/Barcode)
        default_sup_id: Optional[int] = None
        seen_skus: set[str] = set()
        try:
            # si existen productos previos (cuando reset=False)
            for s in sess.query(Product.sku).all():
                if s[0]:
                    seen_skus.add(str(s[0]).strip())
        except Exception:
            pass

        seen_barcodes: set[str] = set()
        try:
            for b in sess.query(Product.barcode).all():
                if b[0]:
                    seen_barcodes.add(str(b[0]).strip())
        except Exception:
            pass

        for r in _read("productos"):
            try:
                # Resolver proveedor: garantizar FK válido
                prov_id = int(r.get("id_proveedor")) if r.get("id_proveedor") is not None else None
                if prov_id is None or sess.get(Supplier, prov_id) is None:
                    if prov_id is None:
                        # Usar/crear proveedor por defecto
                        if default_sup_id is None:
                            try:
                                existing_def = sess.query(Supplier).filter(Supplier.rut == "FAKE-DEFAULT").first()
                            except Exception:
                                existing_def = None
                            if existing_def is None:
                                sup_def = Supplier(razon_social="Proveedor Importado", rut="FAKE-DEFAULT")
                                sess.add(sup_def)
                                sess.flush()
                                default_sup_id = int(sup_def.id)
                            else:
                                default_sup_id = int(existing_def.id)
                        prov_id = default_sup_id
                    else:
                        # Crear stub con el ID esperado
                        sup_stub = Supplier(id=prov_id, razon_social=f"Proveedor {prov_id}", rut=f"FAKE-{prov_id}")
                        sess.add(sup_stub)
                        sess.flush()

                # Resolver ubicación: si no existe, dejar en None
                loc_id = int(r.get("id_ubicacion")) if r.get("id_ubicacion") is not None else None
                if loc_id is not None and loc_id not in existing_locations:
                    loc_id = None

                # SKU único
                raw_sku = str(r.get("sku") or f"SKU-{r.get('id')}")
                base_sku = raw_sku.strip() or f"SKU-{r.get('id')}"
                sku_val = base_sku
                dup_i = 2
                while sku_val in seen_skus:
                    sku_val = f"{base_sku}-{dup_i}"
                    dup_i += 1
                seen_skus.add(sku_val)

                # Barcode único opcional
                bc_val = None
                if "barcode" in r and r.get("barcode"):
                    cand = str(r.get("barcode")).strip()
                    if cand and cand not in seen_barcodes:
                        bc_val = cand
                        seen_barcodes.add(cand)

                prod = Product(
                    id=int(r.get("id")),
                    nombre=str(r.get("nombre") or ""),
                    sku=sku_val,
                    precio_compra=_to_float(r.get("precio_compra")) or 0,
                    precio_venta=_to_float(r.get("precio_venta")) or 0,
                    stock_actual=int(r.get("stock_actual") or 0),
                    unidad_medida=r.get("unidad_medida") or None,
                    familia=r.get("familia") or None,
                    id_proveedor=int(prov_id),
                    id_ubicacion=loc_id,
                    image_path=r.get("image_path") or None,
                    barcode=bc_val,
                )
                sess.add(prod)
            except Exception:
                pass
        sess.flush()

        primary_images_by_product: dict[int, Path] = {}
        for r in _read("imagenes_productos"):
            try:
                pid = int(r.get("id_producto"))
                rel_path = str(r.get("ruta_relativa") or "").strip().replace("\\", "/")
                payload = str(r.get("archivo_base64") or "").strip()
                if not rel_path or not payload:
                    continue
                target = (PRODUCTS_DIR / rel_path).resolve()
                try:
                    target.relative_to(PRODUCTS_DIR.resolve())
                except Exception:
                    continue
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_bytes(base64.b64decode(payload))
                if str(r.get("es_principal") or "").strip() == "1":
                    primary_images_by_product[pid] = target
            except Exception:
                pass

        if primary_images_by_product:
            for pid, target in primary_images_by_product.items():
                try:
                    prod = sess.get(Product, int(pid))
                    if prod is not None:
                        prod.image_path = str(target)
                except Exception:
                    pass

        # Órdenes: compras, ventas y recepciones
        for r in _read("ordenes"):
            try:
                tipo = (str(r.get("tipo") or "").strip().lower())
                if tipo == "compra":
                    pur = Purchase(
                        id=int(r.get("id")),
                        id_proveedor=int(r.get("tercero_id") or r.get("id_proveedor") or 0),
                        fecha_compra=r.get("fecha"),
                        total_compra=_to_val(r.get("total")) or 0,
                        estado=str(r.get("estado") or "Pendiente"),
                        numero_documento=str(r.get("doc_numero") or ""),
                        fecha_documento=r.get("fecha_documento"),
                        fecha_contable=r.get("fecha_contable"),
                        fecha_vencimiento=r.get("fecha_vencimiento"),
                        moneda=str(r.get("moneda") or ""),
                        tasa_cambio=_to_val(r.get("tasa_cambio")) if r.get("tasa_cambio") is not None else None,
                        unidad_negocio=str(r.get("unidad_negocio") or ""),
                        proporcionalidad=str(r.get("proporcionalidad") or ""),
                        atencion=str(r.get("atencion") or ""),
                        tipo_descuento=str(r.get("tipo_descuento") or ""),
                        descuento=_to_val(r.get("descuento")) if r.get("descuento") is not None else None,
                        ajuste_iva=_to_val(r.get("ajuste_iva")) if r.get("ajuste_iva") is not None else None,
                        stock_policy=str(r.get("stock_policy") or ""),
                        referencia=str(r.get("referencia") or ""),
                        ajuste_impuesto=_to_val(r.get("ajuste_impuesto")) if r.get("ajuste_impuesto") is not None else None,
                    )
                    sess.add(pur)
                elif tipo == "venta":
                    sale = Sale(
                        id=int(r.get("ov_id") or r.get("id")),
                        id_cliente=int(r.get("tercero_id") or 0),
                        fecha_venta=r.get("fecha"),
                        total_venta=_to_val(r.get("total")) or 0,
                        estado=str(r.get("estado") or "Confirmada"),
                        numero_documento=str(r.get("numero_documento") or r.get("doc_numero") or ""),
                        mes_referencia=str(r.get("mes_referencia") or ""),
                        monto_neto=_to_val(r.get("monto_neto")) if r.get("monto_neto") is not None else None,
                        monto_iva=_to_val(r.get("monto_iva")) if r.get("monto_iva") is not None else None,
                        fecha_pagado=r.get("fecha_pagado"),
                        nota=str(r.get("nota") or ""),
                        estado_externo=str(r.get("estado_externo") or ""),
                        origen=str(r.get("origen") or ""),
                    )
                    sess.add(sale)
                elif tipo == "recepcion":
                    rec = Reception(
                        id=int(r.get("id")),
                        id_compra=int(r.get("oc_id") or 0),
                        tipo_doc=str(r.get("doc_tipo") or ""),
                        numero_documento=str(r.get("doc_numero") or ""),
                        fecha=r.get("fecha"),
                    )
                    sess.add(rec)
            except Exception:
                pass

        sess.flush()

        for r in _read("detalles_compra"):
            try:
                det = PurchaseDetail(
                    id=int(r.get("id")),
                    id_compra=int(r.get("id_compra")),
                    id_producto=int(r.get("id_producto")),
                    cantidad=int(r.get("cantidad") or 0),
                    received_qty=int(r.get("received_qty") or 0),
                    precio_unitario=_to_float(r.get("precio_unitario")) or 0,
                    subtotal=_to_float(r.get("subtotal")) or 0,
                )
                sess.add(det)
            except Exception:
                pass

        for r in _read("detalles_venta"):
            try:
                det = SaleDetail(
                    id=int(r.get("id")),
                    id_venta=int(r.get("id_venta")),
                    id_producto=int(r.get("id_producto")),
                    cantidad=int(r.get("cantidad") or 0),
                    precio_unitario=_to_float(r.get("precio_unitario")) or 0,
                    subtotal=_to_float(r.get("subtotal")) or 0,
                )
                sess.add(det)
            except Exception:
                pass

        for r in _read("servicios_venta"):
            try:
                det = SaleServiceDetail(
                    id=int(r.get("id")),
                    id_venta=int(r.get("id_venta")),
                    descripcion=str(r.get("descripcion") or "").strip(),
                    cantidad=int(r.get("cantidad") or 0),
                    precio_unitario=_to_float(r.get("precio_unitario")) or 0,
                    subtotal=_to_float(r.get("subtotal")) or 0,
                    afecto_iva=_to_bool(r.get("afecto_iva"), default=True),
                )
                sess.add(det)
            except Exception:
                pass

        for r in _read("movimientos_entrada"):
            try:
                mov = StockEntry(
                    id=int(r.get("id")),
                    id_producto=int(r.get("id_producto")),
                    id_ubicacion=int(r.get("id_ubicacion")) if r.get("id_ubicacion") is not None else None,
                    id_recepcion=int(r.get("id_recepcion")) if r.get("id_recepcion") is not None else None,
                    cantidad=int(r.get("cantidad") or 0),
                    motivo=str(r.get("motivo") or ""),
                    lote=str(r.get("lote") or "") or None,
                    serie=str(r.get("serie") or "") or None,
                    fecha_vencimiento=r.get("fecha_vencimiento"),
                    fecha=r.get("fecha"),
                )
                sess.add(mov)
            except Exception:
                pass

        for r in _read("movimientos_salida"):
            try:
                mov = StockExit(
                    id=int(r.get("id")),
                    id_producto=int(r.get("id_producto")),
                    id_ubicacion=int(r.get("id_ubicacion")) if r.get("id_ubicacion") is not None else None,
                    cantidad=int(r.get("cantidad") or 0),
                    motivo=str(r.get("motivo") or ""),
                    lote=str(r.get("lote") or "") or None,
                    serie=str(r.get("serie") or "") or None,
                    fecha=r.get("fecha"),
                )
                sess.add(mov)
            except Exception:
                pass

        # Snapshot de inventario: sobreescribe stock_actual
        inv = _read("inventario")
        if inv:
            for r in inv:
                try:
                    pid = int(r.get("id_producto"))
                    stk = int(r.get("stock_actual") or 0)
                    prod = sess.get(Product, pid)
                    if prod is not None:
                        prod.stock_actual = stk
                except Exception:
                    pass

        sess.commit()
    except Exception:
        sess.rollback()
        raise
    finally:
        try:
            wb.close()
        except Exception:
            pass

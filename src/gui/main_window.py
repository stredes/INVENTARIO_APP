# src/gui/main_window.py
from __future__ import annotations
import sys
import configparser
from pathlib import Path
import tkinter as tk
from tkinter import ttk, Menu
from tkinter import filedialog, messagebox

from src.gui.products_view import ProductsView
from src.gui.suppliers_view import SuppliersView
from src.gui.customers_view import CustomersView
from src.gui.purchases_view import PurchasesView
from src.gui.sales_view import SalesView
from src.gui.inventory_view import InventoryView
from src.gui.orders_admin_view import OrdersAdminView
from src.reports.report_center import ReportCenter  # â† NUEVO
from src.gui.catalog_view import CatalogView
from src.gui.tutorial_center import TutorialCenter

from src.gui.theme_manager import ThemeManager
from src.gui.widgets.status_bar import StatusBar
from src.gui.widgets.toast import Toast
from src.gui.widgets.command_palette import CommandPalette, CommandAction
from src.utils.printers import (
    get_document_printer,
    set_document_printer,
    get_label_printer,
    set_label_printer,
    print_file_windows,
)

UI_STATE_PATH = Path("config/ui_state.ini")


class MainWindow(ttk.Frame):
    def __init__(self, master: tk.Misc):
        super().__init__(master, padding=10)

        # â¬‡ï¸ NO USAR self._root (colisiona con tk._root())
        self.app_root: tk.Tk = self.winfo_toplevel()

        # Tema y menú
        ThemeManager.attach(self.app_root)
        self._build_menu()
        self._tutorial_window = None
        self._build_top_bar()

        # Notebook + tabs
        self.notebook = ttk.Notebook(self)
        self.products_tab = ProductsView(self.notebook)
        self.suppliers_tab = SuppliersView(self.notebook)
        self.customers_tab = CustomersView(self.notebook)
        self.purchases_tab = PurchasesView(self.notebook)
        # Construcción segura de Ventas con fallback visible si hay error
        try:
            self.sales_tab = SalesView(self.notebook)
        except Exception as ex:
            err = ttk.Frame(self.notebook)
            ttk.Label(err, text=f"No se pudo cargar 'Ventas':\n{ex}", foreground="#b00020").pack(padx=12, pady=12, anchor='w')
            self.sales_tab = err
        self.inventory_tab = InventoryView(self.notebook)
        self.orders_admin_tab = OrdersAdminView(self.notebook)
        self.report_center_tab = ReportCenter(self.notebook)  # â† NUEVO
        self.catalog_tab = CatalogView(self.notebook)

        self.notebook.add(self.products_tab, text="Productos")
        self.notebook.add(self.suppliers_tab, text="Proveedores")
        self.notebook.add(self.customers_tab, text="Clientes")
        self.notebook.add(self.purchases_tab, text="Compras")
        self.notebook.add(self.sales_tab, text="Ventas")
        self.notebook.add(self.inventory_tab, text="Inventario")
        self.notebook.add(self.orders_admin_tab, text="Órdenes")
        self.notebook.add(self.report_center_tab, text="Informes")  # â† NUEVO
        self.notebook.add(self.catalog_tab, text="Catálogo")

        self._tutorial_modules = self._build_tutorial_modules()
        self._tutorial_tab_by_name = {
            "Productos": self.products_tab,
            "Proveedores": self.suppliers_tab,
            "Clientes": self.customers_tab,
            "Compras": self.purchases_tab,
            "Ventas": self.sales_tab,
            "Inventario": self.inventory_tab,
            "Ordenes": self.orders_admin_tab,
            "Informes": self.report_center_tab,
            "Catalogo": self.catalog_tab,
        }
        self._tutorial_name_by_tab = {v: k for k, v in self._tutorial_tab_by_name.items()}

        self.notebook.pack(fill="both", expand=True)
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_change)

        # Status bar
        self.status = StatusBar(self)
        self.status.pack(side="bottom", fill="x")
        self.status.set_message("Listo")
        self.status.set_info(user="admin", env="dev")
        self.status.set_db_status("OK")

        # Atajos
        self._setup_shortcuts()

        # Persistencia
        self._restore_ui_state()
        self.app_root.bind("<Destroy>", self._on_root_destroy, add="+")

        # Toast de ayuda
        Toast.show(self.app_root, "Ctrl+K: Paleta de comandos", kind="info", ms=1800)

    def _generate_catalog(self) -> None:
        """Genera catálogo PDF de productos (imagen, stock y precio sin IVA)."""
        try:
            from src.reports.catalog_generator import generate_products_catalog
            out = generate_products_catalog(self.products_tab.session, auto_open=True)
            Toast.show(self.app_root, f"Catálogo generado: {out}", kind="success")
        except Exception as ex:
            Toast.show(self.app_root, f"Error al generar catálogo: {ex}", kind="danger", position="tr")

    def _open_db_connection_dialog(self) -> None:
        try:
            from src.gui.db_connection_dialog import DBConnectionDialog
        except Exception as ex:
            Toast.show(self.app_root, f"No se pudo cargar el diálogo de DB: {ex}", kind="danger")
            return
        DBConnectionDialog(self)

    def _build_top_bar(self) -> None:
        bar = ttk.Frame(self)
        bar.pack(fill="x", expand=False, pady=(0, 6))
        self._tutorial_btn = ttk.Button(bar, text="?", width=3, command=self._open_tutorial_center)
        self._tutorial_btn.pack(side="right")
        ttk.Label(bar, text="Tutoriales").pack(side="right", padx=(0, 6))

    def _build_menu(self) -> None:
        menubar = Menu(self.app_root)
        self.app_root.config(menu=menubar)

        m_file = Menu(menubar, tearoff=False)
        m_file.add_command(label="Paleta de comandosâ€¦    Ctrl+K", command=self._open_palette)
        m_file.add_separator()
        m_file.add_command(label="Salir", command=self.app_root.quit)
        menubar.add_cascade(label="Archivo", menu=m_file)

        ThemeManager.build_menu(menubar)

        # Menú de Impresoras (Windows/Bluetooth)
        m_prn = Menu(menubar, tearoff=False)
        m_prn.add_command(label="Impresora de documentos…", command=self._choose_document_printer)
        m_prn.add_command(label="Impresora de etiquetas…", command=self._choose_label_printer)
        m_prn.add_separator()
        m_prn.add_command(label="Probar impresión de documento", command=self._test_print_document)
        m_prn.add_command(label="Probar impresión de etiqueta", command=self._test_print_label)
        m_prn.add_separator()
        m_prn.add_command(label="Explorar Bluetooth (experimental)…", command=self._open_bt_scan)
        menubar.add_cascade(label="Impresora", menu=m_prn)

        m_tools = Menu(menubar, tearoff=False)
        m_tools.add_command(label="Nuevo (Ctrl+N)", command=self._new_current)
        m_tools.add_command(label="Guardar (Ctrl+S)", command=self._save_current)
        m_tools.add_command(label="Imprimir (Ctrl+P)", command=self._print_current)
        m_tools.add_command(label="Refrescar aplicacion (F5)", command=self._refresh_all)
        m_tools.add_separator()
        m_tools.add_command(label="Generador de catálogos", command=self._generate_catalog)
        m_tools.add_separator()
        m_tools.add_command(label="Conexión a BDâ€¦", command=self._open_db_connection_dialog)
        # Importacion SQL masiva
        m_tools.add_separator()
        m_tools.add_command(label="Importacion SQL masiva...", command=self._open_sql_importer)
        # Limpiador de base de datos (BORRA TODO)
        m_tools.add_separator()
        m_tools.add_command(label="Limpiar base de datos...", command=self._wipe_databases)
        # ERP Backup
        m_tools.add_separator()
        m_tools.add_command(label="ERP: Exportar backup (XLSX)...", command=self._erp_export_backup)
        m_tools.add_command(label="ERP: Importar backup (XLSX)...", command=self._erp_import_backup)
        # APP Backup (5 hojas)
        m_tools.add_separator()
        m_tools.add_command(label="APP: Exportar backup (XLSX)...", command=self._app_export_backup)
        m_tools.add_command(label="APP: Importar backup (XLSX)...", command=self._app_import_backup)
        menubar.add_cascade(label="Herramientas", menu=m_tools)

        m_view = Menu(menubar, tearoff=False)
        m_view.add_command(label="Paleta de comandos…", command=self._open_palette)
        m_view.add_command(label="Editor de información…", command=self._open_company_editor)
        m_view.add_command(label="Ir a Informes", command=self.show_report_center)
        menubar.add_cascade(label="Ver", menu=m_view)

        m_help = Menu(menubar, tearoff=False)
        m_help.add_command(label="Tutoriales...", command=self._open_tutorial_center)
        menubar.add_cascade(label="Ayuda", menu=m_help)

    # ---------------- Impresoras: seleccionar / probar ---------------- #
    def _choose_document_printer(self) -> None:
        try:
            from src.gui.printer_select_dialog import PrinterSelectDialog
        except Exception as ex:
            try:
                Toast.show(self.app_root, f"No se pudo abrir selector: {ex}", kind="danger")
            except Exception:
                pass
            return
        initial = get_document_printer() or None
        dlg = PrinterSelectDialog(self, initial=initial)
        self.wait_window(dlg)
        if getattr(dlg, "result", None):
            set_document_printer(dlg.result)
            try:
                Toast.show(self.app_root, f"Impresora de documentos: {dlg.result}", kind="success")
            except Exception:
                pass

    def _choose_label_printer(self) -> None:
        try:
            from src.gui.printer_select_dialog import PrinterSelectDialog
        except Exception as ex:
            try:
                Toast.show(self.app_root, f"No se pudo abrir selector: {ex}", kind="danger")
            except Exception:
                pass
            return
        initial = get_label_printer() or None
        dlg = PrinterSelectDialog(self, initial=initial)
        self.wait_window(dlg)
        if getattr(dlg, "result", None):
            set_label_printer(dlg.result)
            try:
                Toast.show(self.app_root, f"Impresora de etiquetas: {dlg.result}", kind="success")
            except Exception:
                pass

    def _test_print_document(self) -> None:
        prn = get_document_printer()
        if not prn:
            try:
                Toast.show(self.app_root, "Configure primero la impresora de documentos.", kind="warning")
            except Exception:
                pass
            return
        # Crear un XLSX mínimo y enviarlo a imprimir
        try:
            from openpyxl import Workbook  # type: ignore
            from pathlib import Path
            from tempfile import gettempdir
            from src.reports.print_backend import print_xlsx
            wb = Workbook()
            ws = wb.active
            ws.title = "Prueba"
            ws.cell(row=1, column=1).value = "Prueba de impresión"
            ws.cell(row=2, column=1).value = "Si ves esta hoja, la integración funciona."
            out = Path(gettempdir()) / "test_print.xlsx"
            wb.save(out)
            print_xlsx(out, printer_name=prn)
            try:
                Toast.show(self.app_root, f"Enviado a '{prn}'", kind="success")
            except Exception:
                pass
        except Exception as ex:
            try:
                Toast.show(self.app_root, f"Fallo prueba: {ex}", kind="danger")
            except Exception:
                pass

    def _test_print_label(self) -> None:
        prn = get_label_printer()
        try:
            from src.reports.barcode_label import generate_label_pdf
            from tempfile import gettempdir
            from pathlib import Path
            pdf = generate_label_pdf("TEST-123", text="Etiqueta de prueba", label_w_mm=50, label_h_mm=30, copies=1, auto_open=False)
            # Intentar envío directo en Windows si hay impresora definida
            try:
                print_file_windows(pdf, printer_name=prn)
                Toast.show(self.app_root, f"Etiqueta enviada a '{prn or 'predeterminada'}'", kind="success")
            except Exception as ex:
                # Abrir para impresión manual
                try:
                    Toast.show(self.app_root, f"No se pudo enviar directo: {ex}. Abriendo visor.", kind="warning")
                except Exception:
                    pass
                try:
                    import webbrowser
                    webbrowser.open(str(pdf))
                except Exception:
                    pass
        except Exception as ex:
            try:
                Toast.show(self.app_root, f"Fallo prueba etiquetas: {ex}", kind="danger")
            except Exception:
                pass

    def _open_bt_scan(self) -> None:
        """Explorador básico de dispositivos Bluetooth (si 'bleak' está disponible)."""
        try:
            from src.gui.bluetooth_scan_dialog import BluetoothScanDialog
            BluetoothScanDialog(self)
        except Exception as ex:
            try:
                Toast.show(self.app_root, f"Bluetooth no disponible: {ex}", kind="warning")
            except Exception:
                pass

    def _setup_shortcuts(self) -> None:
        self.bind_all("<Control-k>", lambda e: self._open_palette())
        self.bind_all("<Control-n>", lambda e: self._new_current())
        self.bind_all("<Control-s>", lambda e: self._save_current())
        self.bind_all("<Control-p>", lambda e: self._print_current())
        # F5 refresca toda la aplicacion
        try:
            self.bind_all("<F5>", lambda e: self._refresh_all())
        except Exception:
            pass

    # ---------------- ERP BACKUP ---------------- #
    def _erp_export_backup(self) -> None:
        try:
            from src.erp.tools.backup import export_erp_to_xlsx
        except Exception as ex:
            Toast.show(self.app_root, f"No se pudo cargar backup: {ex}", kind="danger")
            return
        # Seleccionar destino
        try:
            initial = Path.home() / "Downloads" / f"ERP_backup.xlsx"
        except Exception:
            initial = None
        path = filedialog.asksaveasfilename(
            parent=self.app_root,
            title="Guardar backup ERP",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=(initial.name if initial else None),
            initialdir=(str(initial.parent) if initial else None),
        )
        if not path:
            return
        try:
            out = export_erp_to_xlsx(Path(path), auto_open=True)
            Toast.show(self.app_root, f"Backup generado: {out}", kind="success")
        except Exception as ex:
            Toast.show(self.app_root, f"Error al generar backup: {ex}", kind="danger")

    def _erp_import_backup(self) -> None:
        try:
            from src.erp.tools.backup import import_erp_from_xlsx
        except Exception as ex:
            Toast.show(self.app_root, f"No se pudo cargar backup: {ex}", kind="danger")
            return
        path = filedialog.askopenfilename(
            parent=self.app_root,
            title="Seleccionar backup ERP (XLSX)",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if not path:
            return
        if not messagebox.askyesno("Confirmar", "Esto reemplazará los datos ERP por el backup seleccionado. ¿Continuar?", parent=self.app_root):
            return
        try:
            import_erp_from_xlsx(Path(path))
            Toast.show(self.app_root, "Backup importado correctamente", kind="success")
        except Exception as ex:
            Toast.show(self.app_root, f"Error al importar backup: {ex}", kind="danger")

    # ---------------- APP BACKUP (5 hojas) ---------------- #
    def _app_export_backup(self) -> None:
        try:
            from src.erp.tools.backup import export_app_backup_to_xlsx
        except Exception as ex:
            try:
                Toast.show(self.app_root, f"No se pudo cargar backup APP: {ex}", kind="danger")
            except Exception:
                pass
            return
        # Seleccionar destino
        try:
            initial = Path.home() / "Downloads" / f"APP_backup.xlsx"
        except Exception:
            initial = None
        path = filedialog.asksaveasfilename(
            parent=self.app_root,
            title="Guardar backup APP",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=(initial.name if initial else None),
            initialdir=(str(initial.parent) if initial else None),
        )
        if not path:
            return
        try:
            out = export_app_backup_to_xlsx(Path(path), auto_open=True)
            Toast.show(self.app_root, f"Backup APP generado: {out}", kind="success")
        except Exception as ex:
            Toast.show(self.app_root, f"Error al exportar backup APP: {ex}", kind="danger")

    def _app_import_backup(self) -> None:
        try:
            from src.erp.tools.backup import import_app_backup_from_xlsx
        except Exception as ex:
            try:
                Toast.show(self.app_root, f"No se pudo cargar importador APP: {ex}", kind="danger")
            except Exception:
                pass
            return
        path = filedialog.askopenfilename(
            parent=self.app_root,
            title="Seleccionar backup APP (XLSX)",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if not path:
            return
        if not messagebox.askyesno(
            "Confirmar",
            "Esto reemplazará productos, clientes, proveedores, órdenes y stock actual por el archivo seleccionado.\n\n¿Desea continuar?",
            parent=self.app_root,
        ):
            return
        try:
            import_app_backup_from_xlsx(Path(path), reset=True)
            Toast.show(self.app_root, "Backup APP importado correctamente", kind="success")
            # Refrescar vistas para reflejar los datos importados
            self._refresh_all()
        except Exception as ex:
            Toast.show(self.app_root, f"Error al importar backup APP: {ex}", kind="danger")

    def _open_palette(self) -> None:
        actions = self._build_actions()
        CommandPalette.open(self.app_root, actions=actions)

    def _build_actions(self) -> list[CommandAction]:
        view = self._current_view()
        actions: list[CommandAction] = [
            CommandAction("open_tutorials", "Tutoriales", callback=self._open_tutorial_center, keywords=["ayuda", "tutorial"]),
            CommandAction("go_products", "Ir a Productos", callback=self.show_products, keywords=["inventario", "stock"]),
            CommandAction("go_suppliers", "Ir a Proveedores", callback=self.show_suppliers, keywords=["proveedores"]),
            CommandAction("go_customers", "Ir a Clientes", callback=self.show_customers, keywords=["clientes"]),
            CommandAction("go_purchases", "Ir a Compras", callback=self.show_purchases, keywords=["oc", "orden de compra"]),
            CommandAction("go_sales", "Ir a Ventas", callback=self.show_sales, keywords=["ov", "boletas", "ventas"]),
            CommandAction("go_inventory", "Ir a Inventario", callback=self.show_inventory, keywords=["kardex", "bodega"]),
            CommandAction("go_orders", "Ir a Ã“rdenes", callback=self.show_orders_admin, keywords=["admin", "oc", "ov"]),
            CommandAction("go_reports", "Ir a Informes", callback=self.show_report_center, keywords=["reportes", "informes"]),  # â† NUEVO
            CommandAction("go_catalog", "Ir a Catálogo", callback=self.show_catalog, keywords=["catalogo", "pdf"]),
            CommandAction("act_new", "Nuevo registro", callback=self._new_current, category="Acción", shortcut="Ctrl+N"),
            CommandAction("act_save", "Guardar cambios", callback=self._save_current, category="Acción", shortcut="Ctrl+S"),
            CommandAction("act_print", "Imprimir vista", callback=self._print_current, category="Acción", shortcut="Ctrl+P"),
        ]
        if hasattr(view, "actions") and callable(getattr(view, "actions")):
            try:
                for label, func in view.actions():
                    actions.append(CommandAction(f"view_{label}", label, callback=func, category="Vista"))
            except Exception:
                pass
        return actions

    def _build_tutorial_modules(self) -> dict[str, list[str]]:
        return {
            "Productos": [
                "Filtra la grilla por ID, codigo o nombre para ubicar productos.",
                "Completa el formulario con nombre, SKU, proveedor y unidad.",
                "Define precio compra, IVA y margen para recalcular precio venta.",
                "Carga imagen o imprime etiquetas desde el panel de codigo de barras.",
                "Guarda; doble clic en una fila para editar o actualizar.",
            ],
            "Proveedores": [
                "Registra razon social, RUT y datos de contacto.",
                "Usa Agregar para crear un proveedor nuevo.",
                "Doble clic en una fila para editar y Guardar cambios.",
                "Elimina registros o limpia el formulario cuando termines.",
            ],
            "Clientes": [
                "Registra razon social, RUT y datos de contacto.",
                "Usa Agregar para crear clientes nuevos.",
                "Doble clic en una fila para editar y Guardar cambios.",
                "Elimina registros o limpia el formulario cuando termines.",
            ],
            "Compras": [
                "Selecciona proveedor y agrega items en el detalle.",
                "Define estado, forma de pago y si suma stock.",
                "Modo Compra: confirma compra y genera OC o cotizacion.",
                "Modo Recepcion: vincula OC y documento (factura o guia).",
                "Confirma recepcion para sumar stock y dejar historial.",
            ],
            "Ventas": [
                "Selecciona cliente y agrega productos con cantidad y precio.",
                "Define estado y pago; activa descuento de stock si aplica.",
                "Genera OV o cotizacion desde los botones del modulo.",
                "Modo cajero: escanea SKU y cobra rapido con F12.",
                "Confirma para registrar la venta y actualizar el stock.",
            ],
            "Inventario": [
                "Revisa stock en la grilla y usa filtros para acotar.",
                "Configura min/max para resaltar niveles criticos.",
                "Imprime reportes o exporta XLSX cuando lo necesites.",
                "Selecciona un producto para ver/imprimir codigo de barras.",
                "Ajusta el auto refresco segun tu ritmo de trabajo.",
            ],
            "Ordenes": [
                "Usa las pestañas para ver todas, compras, ventas y recepciones.",
                "Doble clic abre el detalle de una orden o recepcion.",
                "Cambia estados para confirmar, completar o cancelar.",
                "En recepciones revisa documentos vinculados a la OC.",
            ],
            "Informes": [
                "Elige el tipo de informe en el selector superior.",
                "Configura filtros de fecha, estado, tercero o producto.",
                "Ejecuta el informe y revisa los resultados en la tabla.",
                "Exporta el reporte para compartir o archivar.",
            ],
            "Catalogo": [
                "Define copias, IVA, diseno, titulo y familia.",
                "Elige que campos mostrar (empresa, SKU, stock, precios).",
                "Genera una vista previa para validar el diseno.",
                "Imprime o guarda el PDF del catalogo.",
            ],
        }

    def _open_tutorial_center(self) -> None:
        try:
            if self._tutorial_window and self._tutorial_window.winfo_exists():
                self._tutorial_window.lift()
                self._tutorial_window.focus_force()
                return
        except Exception:
            pass

        start_module = self._current_tutorial_module()
        self._tutorial_window = TutorialCenter(
            self.app_root,
            modules=self._tutorial_modules,
            start_module=start_module,
            on_open_module=self._open_tutorial_module,
        )
        try:
            self._tutorial_window.bind("<Destroy>", self._on_tutorial_destroy, add="+")
        except Exception:
            pass

    def _on_tutorial_destroy(self, event) -> None:
        if event.widget is self._tutorial_window:
            self._tutorial_window = None

    def _current_tutorial_module(self) -> str | None:
        view = self._current_view()
        return self._tutorial_name_by_tab.get(view)

    def _open_tutorial_module(self, name: str) -> None:
        widget = self._tutorial_tab_by_name.get(name)
        if widget is not None:
            self._select_tab_by_widget(widget)

    def _current_view(self) -> ttk.Frame:
        sel = self.notebook.select()
        return self.notebook.nametowidget(sel)

    def _select_tab_by_widget(self, widget: ttk.Frame) -> None:
        self.notebook.select(widget)

    def show_products(self) -> None: self._select_tab_by_widget(self.products_tab)
    def show_suppliers(self) -> None: self._select_tab_by_widget(self.suppliers_tab)
    def show_customers(self) -> None: self._select_tab_by_widget(self.customers_tab)
    def show_purchases(self) -> None: self._select_tab_by_widget(self.purchases_tab)
    def show_sales(self) -> None: self._select_tab_by_widget(self.sales_tab)
    def show_inventory(self) -> None: self._select_tab_by_widget(self.inventory_tab)
    def show_orders_admin(self) -> None: self._select_tab_by_widget(self.orders_admin_tab)
    def show_report_center(self) -> None: self._select_tab_by_widget(self.report_center_tab)  # â† NUEVO
    def show_catalog(self) -> None: self._select_tab_by_widget(self.catalog_tab)

    def _new_current(self) -> None:
        view = self._current_view()
        for name in ("new", "create_new", "new_record"):
            if hasattr(view, name) and callable(getattr(view, name)):
                getattr(view, name)()
                self.status.flash("Nuevo registro", kind="info", ms=1200)
                return
        Toast.show(self.app_root, "La vista actual no soporta 'Nuevo'", kind="warning")

    def _save_current(self) -> None:
        view = self._current_view()
        for name in ("save", "save_changes", "commit"):
            if hasattr(view, name) and callable(getattr(view, name)):
                try:
                    getattr(view, name)()
                    self.status.flash("Guardado con éxito", kind="success", ms=1400)
                except Exception as ex:
                    Toast.show(self.app_root, f"Error al guardar: {ex}", kind="danger", position="tr")
                    self.status.flash("Error al guardar", kind="danger", ms=1400)
                return
        Toast.show(self.app_root, "La vista actual no soporta 'Guardar'", kind="warning")

    def _print_current(self) -> None:
        view = self._current_view()
        for name in ("print_current", "print_inventory", "print", "export_pdf"):  # â† incluye print_inventory
            if hasattr(view, name) and callable(getattr(view, name)):
                try:
                    self.status.progress_start(indeterminate=True)
                    getattr(view, name)()
                    self.status.progress_hide()
                    self.status.flash("Impresión enviada", kind="success", ms=1400)
                except Exception as ex:
                    self.status.progress_hide()
                    Toast.show(self.app_root, f"Error al imprimir: {ex}", kind="danger", position="tr")
                    self.status.flash("Error al imprimir", kind="danger", ms=1400)
                return
        Toast.show(self.app_root, "La vista actual no soporta 'Imprimir'", kind="warning")

    def _on_tab_change(self, _event=None):
        w = self._current_view()
        if hasattr(w, "refresh_lookups"):
            try:
                w.refresh_lookups()
            except Exception as ex:
                # Evita fallar si la vista aún no terminó de construir widgets
                print(f"Error al refrescar lookups: {ex}", file=sys.stderr)
                try:
                    self.app_root.after(200, lambda: getattr(w, "refresh_lookups", lambda: None)())
                except Exception:
                    pass
        tab_text = self.notebook.tab(self.notebook.select(), "text")
        self.status.set_message(f"Vista: {tab_text} â€” Ctrl+K para comandos")
        self._save_last_tab_index()

    def _restore_ui_state(self) -> None:
        cfg = configparser.ConfigParser(strict=False)
        if UI_STATE_PATH.exists():
            cfg.read(UI_STATE_PATH, encoding="utf-8")
        geom = cfg.get("mainwindow", "geometry", fallback="")
        if geom:
            # Restaura la geometría, pero asegúrate de que quede visible
            try:
                self.app_root.geometry(geom)
            except Exception:
                pass
            self._ensure_on_screen()
        else:
            # Si no hay geometría guardada, centra dentro de la pantalla actual
            self._center_on_screen()
        last_idx = cfg.getint("mainwindow", "last_tab_index", fallback=0)
        try:
            self.notebook.select(last_idx)
        except Exception:
            pass

    def _ensure_on_screen(self) -> None:
        """Garantiza que la ventana esté dentro de los límites visibles.
        Si la geometría previa estaba en un monitor secundario ausente,
        reubica/clampa la ventana para que sea visible.
        """
        root = self.app_root
        try:
            root.update_idletasks()
            # Tamaños actuales (si aún no se calculan, usa requeridos)
            w = root.winfo_width() or root.winfo_reqwidth()
            h = root.winfo_height() or root.winfo_reqheight()
            # Posición actual
            x = root.winfo_x()
            y = root.winfo_y()
            # Tamaño de escritorio virtual (múltiples monitores)
            v_w = max(getattr(root, 'winfo_vrootwidth', lambda: 0)() or 0, root.winfo_screenwidth())
            v_h = max(getattr(root, 'winfo_vrootheight', lambda: 0)() or 0, root.winfo_screenheight())

            # Clamps seguros
            w = min(w, v_w)
            h = min(h, v_h)
            new_x = min(max(x, 0), max(v_w - w, 0))
            new_y = min(max(y, 0), max(v_h - h, 0))

            if new_x != x or new_y != y:
                root.geometry(f"{w}x{h}+{new_x}+{new_y}")
        except Exception:
            # Ante cualquier error, como último recurso centra
            self._center_on_screen()

    def _center_on_screen(self) -> None:
        """Centra la ventana en la pantalla principal actual."""
        root = self.app_root
        try:
            root.update_idletasks()
            w = root.winfo_width() or root.winfo_reqwidth()
            h = root.winfo_height() or root.winfo_reqheight()
            sw = root.winfo_screenwidth()
            sh = root.winfo_screenheight()
            x = max(0, int((sw - w) / 2))
            y = max(0, int((sh - h) / 2))
            root.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            pass

    def _save_last_tab_index(self) -> None:
        idx = self._safe_current_index()
        cfg = configparser.ConfigParser(strict=False)
        if UI_STATE_PATH.exists():
            cfg.read(UI_STATE_PATH, encoding="utf-8")
        cfg.setdefault("mainwindow", {})
        cfg["mainwindow"]["last_tab_index"] = str(idx)
        UI_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
        with UI_STATE_PATH.open("w", encoding="utf-8") as f:
            cfg.write(f)

    def _on_root_destroy(self, event) -> None:
        if event.widget is not self.app_root:
            return
        try:
            geom = self.app_root.winfo_geometry()
            cfg = configparser.ConfigParser(strict=False)
            if UI_STATE_PATH.exists():
                cfg.read(UI_STATE_PATH, encoding="utf-8")
            cfg.setdefault("mainwindow", {})
            cfg["mainwindow"]["geometry"] = geom
            cfg["mainwindow"]["last_tab_index"] = str(self._safe_current_index())
            UI_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
            with UI_STATE_PATH.open("w", encoding="utf-8") as f:
                cfg.write(f)
        except Exception:
            pass

    def _safe_current_index(self) -> int:
        try:
            return self.notebook.index(self.notebook.select())
        except Exception:
            return 0

    def _wipe_databases(self) -> None:
        """Elimina TODOS los datos de la BD ORM y del ERP simple. IRREVERSIBLE."""
        try:
            from tkinter import messagebox
        except Exception:
            return
        if not messagebox.askyesno(
            "Limpiar base de datos",
            "Esto eliminará TODOS los registros (productos, proveedores, clientes,\n"
            "ventas, compras, inventario y documentos ERP).\n\n¿Desea continuar?",
            parent=self.app_root,
        ):
            return

        # 1) Limpiar ORM (SQLAlchemy)
        try:
            from sqlalchemy.orm import Session as _S
            from src.data.models import (
                SaleDetail, Sale, PurchaseDetail, Purchase,
                StockEntry, StockExit, Product, Supplier, Customer, Location,
                Reception,
            )
            sess: _S = self.products_tab.session
            # Orden seguro respetando FKs
            # 1) Movimientos primero (dependen de productos y recepciones)
            sess.query(StockEntry).delete()
            sess.query(StockExit).delete()
            # 2) Detalles
            sess.query(SaleDetail).delete()
            sess.query(PurchaseDetail).delete()
            # 3) Cabeceras hijas (recepciones) antes de compras
            try:
                sess.query(Reception).delete()
            except Exception:
                pass
            sess.query(Sale).delete()
            sess.query(Purchase).delete()
            sess.query(Product).delete()
            sess.query(Supplier).delete()
            sess.query(Customer).delete()
            sess.query(Location).delete()
            sess.commit()
        except Exception as ex:
            try:
                Toast.show(self.app_root, f"Error limpiando ORM: {ex}", kind="danger", position="tr")
            except Exception:
                pass
            return

        # 2) Limpiar ERP (sqlite simple) si existe
        try:
            from src.erp.core.database import get_connection, DEFAULT_DB_PATH
            conn = get_connection(DEFAULT_DB_PATH)
            try:
                conn.execute("PRAGMA foreign_keys = ON;")
                conn.execute("DELETE FROM detalles;")
                conn.execute("DELETE FROM documentos;")
                conn.execute("DELETE FROM log_auditoria;")
                conn.commit()
            finally:
                conn.close()
        except Exception:
            # Si no existe o falla, lo omitimos silenciosamente
            pass

        # 3) Notificar
        try:
            Toast.show(self.app_root, "Base de datos limpiada", kind="success")
        except Exception:
            messagebox.showinfo("BD", "Base de datos limpiada")

    def _seed_surt_fake(self) -> None:
        """Seed SURT VENTAS (fake): proveedor + productos, opcional compra/stock."""
        try:
            from tkinter import messagebox
            from scripts import seed_surt_ventas as seed
        except Exception as ex:
            try:
                Toast.show(self.app_root, f"No se pudo cargar seed: {ex}", kind="danger")
            except Exception:
                pass
            return

        if not messagebox.askyesno(
            "Sembrar SURT",
            "Se insertaran/actualizaran proveedor y productos basicos.\n\n¿Continuar?",
            parent=self.app_root,
        ):
            return

        try:
            create_purchase = messagebox.askyesno(
                "Sembrar SURT",
                "¿Crear tambien una Orden de Compra (Pendiente)?",
                parent=self.app_root,
            )
            sum_stock = messagebox.askyesno(
                "Sembrar SURT",
                "¿Sumar stock segun cantidades de la orden?",
                parent=self.app_root,
            )

            session = self.products_tab.session
            supplier = seed.ensure_supplier(session, **seed.SUPPLIER)
            seed.upsert_products(session, supplier, seed.ITEMS, margen_pct=30.0)
            if create_purchase:
                seed.create_purchase(session, supplier, seed.ITEMS, estado="Pendiente")
            if sum_stock:
                seed.sum_stock(session, seed.ITEMS)

            Toast.show(self.app_root, "Seed SURT completado", kind="success")
        except Exception as ex:
            Toast.show(self.app_root, f"Fallo seed SURT: {ex}", kind="danger", position="tr")
    
    def _open_sql_importer(self) -> None:
        try:
            from src.gui.sql_importer_dialog import SqlMassImporter
            SqlMassImporter(self)
        except Exception as ex:
            try:
                Toast.show(self.app_root, f"No se pudo abrir importador SQL: {ex}", kind="danger")
            except Exception:
                pass

    def _refresh_all(self) -> None:
        """Refresca todas las pestañas de la aplicacion de forma segura."""
        views = [
            getattr(self, "products_tab", None),
            getattr(self, "suppliers_tab", None),
            getattr(self, "customers_tab", None),
            getattr(self, "purchases_tab", None),
            getattr(self, "sales_tab", None),
            getattr(self, "inventory_tab", None),
            getattr(self, "orders_admin_tab", None),
            getattr(self, "report_center_tab", None),
            getattr(self, "catalog_tab", None),
        ]
        for v in views:
            if v is None:
                continue
            for name in ("refresh_all", "refresh", "refresh_table", "refresh_lookups"):
                try:
                    fn = getattr(v, name, None)
                    if callable(fn):
                        fn()
                        break
                except Exception:
                    pass
        try:
            self.status.flash("Aplicación refrescada", kind="info", ms=1200)
        except Exception:
            pass







    def _open_company_editor(self) -> None:
        try:
            from src.gui.company_info_editor import CompanyInfoEditor
            CompanyInfoEditor(self)
        except Exception as ex:
            try:
                Toast.show(self.app_root, f"No se pudo abrir Editor de información: {ex}", kind="danger")
            except Exception:
                messagebox.showerror("Error", f"No se pudo abrir Editor de información:\n{ex}")

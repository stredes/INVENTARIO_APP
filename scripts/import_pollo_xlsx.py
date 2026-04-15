from __future__ import annotations

import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.data.database import get_session
from src.data.models import Customer, Sale, SaleServiceDetail
from src.core.sales_manager import SalesManager, ManualSaleItem


def parse_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("$", "").replace(".", "").replace(",", ".")
    return Decimal(text or "0")


def normalize_state(raw: str) -> str:
    state = (raw or "").strip().lower()
    if state == "cancelado":
        return "Pagada"
    if state == "pendiente":
        return "Pendiente"
    if state == "confirmada":
        return "Confirmada"
    return "Pendiente"


def build_placeholder_rut(name: str, taken: set[str]) -> str:
    base = "".join(ch for ch in (name or "cliente").upper() if ch.isalnum())[:10] or "CLIENTE"
    idx = 1
    while True:
        rut = f"AUTO-{base}-{idx}"
        if rut not in taken:
            taken.add(rut)
            return rut
        idx += 1


def ensure_customer(session, name: str, taken_ruts: set[str]) -> Customer:
    customer = session.query(Customer).filter(Customer.razon_social == name).first()
    if customer:
        if customer.rut:
            taken_ruts.add(customer.rut)
        return customer
    customer = Customer(
        razon_social=name,
        rut=build_placeholder_rut(name, taken_ruts),
        contacto=None,
        telefono=None,
        email=None,
        direccion=None,
    )
    session.add(customer)
    session.flush()
    return customer


def find_existing_sale(session, customer_id: int, fecha: datetime, total: Decimal, invoice_number: str):
    by_doc = (
        session.query(Sale)
        .filter(Sale.id_cliente == customer_id)
        .filter(Sale.fecha_venta == fecha)
        .filter(Sale.total_venta == total)
        .filter(Sale.numero_documento == str(invoice_number))
        .first()
    )
    if by_doc:
        return by_doc
    by_desc = (
        session.query(Sale)
        .join(SaleServiceDetail, SaleServiceDetail.id_venta == Sale.id)
        .filter(Sale.id_cliente == customer_id)
        .filter(Sale.fecha_venta == fecha)
        .filter(Sale.total_venta == total)
        .filter(SaleServiceDetail.descripcion.ilike(f"Factura {invoice_number}%"))
        .first()
    )
    return by_desc


def import_xlsx(path: Path) -> dict[str, int]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    session = get_session()
    sm = SalesManager(session)
    taken_ruts = {c.rut for c in session.query(Customer).all() if getattr(c, "rut", None)}
    created_customers = 0
    imported_sales = 0
    updated_sales = 0
    skipped_rows = 0

    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            invoice_number = row[1]
            invoice_date = row[3]
            customer_name = (row[4] or "").strip()
            net_amount = parse_decimal(row[5])
            iva_amount = parse_decimal(row[6])
            total = parse_decimal(row[7])
            raw_status = str(row[8] or "").strip()
            status = normalize_state(raw_status)
            paid_date = row[9]
            note = (row[10] or "").strip()
            month_ref = (row[2] or "").strip()

            if not invoice_number or not invoice_date or not customer_name or total <= 0:
                skipped_rows += 1
                continue

            customer_before = session.query(Customer.id).filter(Customer.razon_social == customer_name).first()
            customer = ensure_customer(session, customer_name, taken_ruts)
            if customer_before is None:
                created_customers += 1

            descripcion = f"Factura {invoice_number}"
            if note:
                descripcion = f"{descripcion} - {note}"

            fecha = invoice_date if isinstance(invoice_date, datetime) else datetime.combine(invoice_date, datetime.min.time())
            fecha_pagado = None
            if paid_date:
                fecha_pagado = paid_date if isinstance(paid_date, datetime) else datetime.combine(paid_date, datetime.min.time())

            existing_sale = find_existing_sale(session, customer.id, fecha, total, str(invoice_number))
            if existing_sale:
                existing_sale.numero_documento = str(invoice_number)
                existing_sale.mes_referencia = month_ref or None
                existing_sale.monto_neto = net_amount
                existing_sale.monto_iva = iva_amount
                existing_sale.fecha_pagado = fecha_pagado
                existing_sale.nota = note or None
                existing_sale.estado_externo = raw_status or None
                existing_sale.origen = "pollo.xlsx"
                existing_sale.estado = status
                det = (
                    session.query(SaleServiceDetail)
                    .filter(SaleServiceDetail.id_venta == existing_sale.id)
                    .order_by(SaleServiceDetail.id.asc())
                    .first()
                )
                if det:
                    det.descripcion = descripcion
                    det.precio_unitario = total
                    det.subtotal = total
                session.commit()
                updated_sales += 1
                continue

            sm.create_sale(
                customer_id=customer.id,
                items=[],
                service_items=[
                    ManualSaleItem(
                        descripcion=descripcion,
                        cantidad=1,
                        precio_unitario=total,
                        afecto_iva=True,
                    )
                ],
                fecha=fecha,
                estado=status,
                apply_to_stock=False,
                numero_documento=str(invoice_number),
                mes_referencia=month_ref or None,
                monto_neto=net_amount,
                monto_iva=iva_amount,
                fecha_pagado=fecha_pagado,
                nota=note or None,
                estado_externo=raw_status or None,
                origen="pollo.xlsx",
            )
            imported_sales += 1

        return {
            "customers_created": created_customers,
            "sales_imported": imported_sales,
            "sales_updated": updated_sales,
            "rows_skipped": skipped_rows,
        }
    finally:
        session.close()


def main() -> int:
    if len(sys.argv) < 2:
        print("Uso: python scripts/import_pollo_xlsx.py <archivo.xlsx>")
        return 1
    path = Path(sys.argv[1]).expanduser()
    if not path.exists():
        print(f"No existe el archivo: {path}")
        return 1
    stats = import_xlsx(path)
    print(
        "Importacion completada:",
        f"clientes nuevos={stats['customers_created']},",
        f"ventas importadas={stats['sales_imported']},",
        f"ventas actualizadas={stats['sales_updated']},",
        f"filas omitidas={stats['rows_skipped']}",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

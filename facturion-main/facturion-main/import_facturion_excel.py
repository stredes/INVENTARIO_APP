from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from database.connection import clear_operational_data, get_connection, initialize_database
from services.report_service import ReportService
from services.settings_service import SettingsService


SPANISH_MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def _to_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def _as_date(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    return None


def _header_map(sheet_name: str, row: tuple[object, ...]) -> dict[str, int]:
    return {str(value).strip().lower(): index for index, value in enumerate(row) if value not in (None, "")}


def _normalize_invoice_rows(workbook_path: Path) -> tuple[list[dict[str, object]], dict[int, set[int]]]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook["FACTURAS"]
    rows: list[dict[str, object]] = []
    month_years: dict[int, set[int]] = defaultdict(set)
    headers = _header_map("FACTURAS", next(sheet.iter_rows(min_row=3, max_row=3, values_only=True)))

    for raw in sheet.iter_rows(min_row=4, values_only=True):
        invoice_date = _as_date(raw[headers["fecha factura"]])
        invoice_number = raw[headers["factura"]]

        if not invoice_date or invoice_number in (None, ""):
            continue

        net_amount = _to_float(raw[headers["neto"]])
        vat_amount = _to_float(raw[headers["iva"]])
        total_amount = _to_float(raw[headers["total"]])
        tag_amount = _to_float(raw[headers["tag"]])
        accountant_amount = _to_float(raw[headers["contador $ 50,000"]])
        deposit_amount = _to_float(raw[headers["monto depositado"]])
        savings_amount = _to_float(raw[headers["ahorro"]])
        deposit_manuel_amount = _to_float(raw[headers["deposito manuel"]])
        paid_vat_amount = _to_float(raw[headers["pago de iva"]])
        paid_accountant_amount = _to_float(raw[headers["pago contador"]])
        paid_savings_amount = _to_float(raw[headers["pago de ahorro"]])
        paid_tag_amount = _to_float(raw[headers["pago tag"]])

        if net_amount <= 0 and vat_amount <= 0 and total_amount <= 0:
            continue

        rows.append(
            {
                "invoice_number": str(invoice_number).strip(),
                "invoice_date": invoice_date.strftime("%Y-%m-%d"),
                "net_amount": ReportService.round_money(net_amount),
                "vat_amount": ReportService.round_money(vat_amount),
                "tag_amount": ReportService.round_money(tag_amount),
                "accountant_amount": ReportService.round_money(accountant_amount),
                "total_amount": ReportService.round_money(total_amount),
                "deposit_amount": ReportService.round_money(deposit_amount),
                "savings_amount": ReportService.round_money(savings_amount),
                "deposit_manuel_amount": ReportService.round_money(deposit_manuel_amount),
                "paid_vat_amount": ReportService.round_money(paid_vat_amount),
                "paid_accountant_amount": ReportService.round_money(paid_accountant_amount),
                "paid_savings_amount": ReportService.round_money(paid_savings_amount),
                "paid_tag_amount": ReportService.round_money(paid_tag_amount),
            }
        )
        month_years[invoice_date.month].add(invoice_date.year)

    return rows, month_years


def _normalize_factura_payment_rows(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook["FACTURAS"]
    headers = _header_map("FACTURAS", next(sheet.iter_rows(min_row=3, max_row=3, values_only=True)))
    grouped: dict[str, dict[str, object]] = {}

    for raw in sheet.iter_rows(min_row=4, values_only=True):
        invoice_date = _as_date(raw[headers["fecha factura"]])
        invoice_number = raw[headers["factura"]]
        if not invoice_date or invoice_number in (None, ""):
            continue

        month_key = invoice_date.strftime("%Y-%m")
        row = grouped.setdefault(
            month_key,
            {
                "month": month_key,
                "sii_vat_amount": 0.0,
                "actual_tag_paid": 0.0,
                "actual_accountant_paid": 0.0,
                "actual_savings_paid": 0.0,
            },
        )
        row["sii_vat_amount"] = ReportService.round_money(
            float(row["sii_vat_amount"]) + _to_float(raw[headers["pago de iva"]])
        )
        row["actual_accountant_paid"] = ReportService.round_money(
            float(row["actual_accountant_paid"]) + _to_float(raw[headers["pago contador"]])
        )
        row["actual_tag_paid"] = ReportService.round_money(
            float(row["actual_tag_paid"]) + _to_float(raw[headers["pago tag"]])
        )
        row["actual_savings_paid"] = ReportService.round_money(
            float(row["actual_savings_paid"]) + _to_float(raw[headers["pago de ahorro"]])
        )

    return [
        row
        for row in grouped.values()
        if float(row["sii_vat_amount"]) > 0
        or float(row["actual_tag_paid"]) > 0
        or float(row["actual_accountant_paid"]) > 0
        or float(row["actual_savings_paid"]) > 0
    ]


def _resolve_payment_month(month_name: str, month_years: dict[int, set[int]], fallback_year: int) -> str | None:
    month_number = SPANISH_MONTHS.get((month_name or "").strip().lower())
    if not month_number:
        return None

    years = sorted(month_years.get(month_number, set()))
    if len(years) == 1:
        year = years[0]
    elif len(years) > 1:
        year = years[-1]
    else:
        year = fallback_year

    return f"{year}-{month_number:02d}"


def _normalize_payment_rows(workbook_path: Path, month_years: dict[int, set[int]]) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook["Ivas"]
    fallback_year = max((year for years in month_years.values() for year in years), default=datetime.today().year)
    rows: list[dict[str, object]] = []

    for raw in sheet.iter_rows(min_row=4, values_only=True):
        month_name = raw[1] if len(raw) > 1 else None
        if not month_name:
            continue

        month_key = _resolve_payment_month(str(month_name), month_years, fallback_year)
        if not month_key:
            continue

        sii_vat_amount = _to_float(raw[2] if len(raw) > 2 else 0)
        actual_tag_paid = _to_float(raw[3] if len(raw) > 3 else 0)
        actual_accountant_paid = _to_float(raw[4] if len(raw) > 4 else 0)
        total_averes = _to_float(raw[5] if len(raw) > 5 else 0)

        if sii_vat_amount <= 0 and actual_tag_paid <= 0 and actual_accountant_paid <= 0 and total_averes <= 0:
            continue

        if sii_vat_amount <= 0 and total_averes > 0:
            sii_vat_amount = total_averes

        rows.append(
            {
                "month": month_key,
                "sii_vat_amount": ReportService.round_money(sii_vat_amount),
                "actual_tag_paid": ReportService.round_money(actual_tag_paid),
                "actual_accountant_paid": ReportService.round_money(actual_accountant_paid),
                "actual_savings_paid": 0.0,
            }
        )

    return rows


def import_workbook(workbook_path: Path, client_name: str, cleanup: bool) -> tuple[int, int, int, int]:
    invoice_rows, month_years = _normalize_invoice_rows(workbook_path)
    payment_rows = _normalize_factura_payment_rows(workbook_path) or _normalize_payment_rows(workbook_path, month_years)
    vat_rate = SettingsService.get_vat_rate()
    description = f"Importado desde {workbook_path.name}"

    initialize_database()
    if cleanup:
        clear_operational_data()

    created_invoices = 0
    updated_invoices = 0
    created_payments = 0
    updated_payments = 0

    with get_connection() as connection:
        for row in invoice_rows:
            totals = ReportService.calculate_invoice_totals(
                net_amount=float(row["net_amount"]),
                vat_rate=vat_rate,
                tag_amount=float(row["tag_amount"]),
                accountant_amount=float(row["accountant_amount"]),
            )
            existing = connection.execute(
                "SELECT id FROM invoices WHERE invoice_number = ? AND invoice_date = ?",
                (row["invoice_number"], row["invoice_date"]),
            ).fetchone()

            payload = (
                row["invoice_number"],
                row["invoice_date"],
                client_name,
                description,
                float(row["net_amount"]),
                vat_rate,
                float(row["vat_amount"]) or totals["vat_amount"],
                float(row["tag_amount"]),
                float(row["accountant_amount"]),
                float(row["total_amount"]) or totals["total_amount"],
                float(row["deposit_amount"]),
                float(row["savings_amount"]),
                float(row["deposit_manuel_amount"]),
                float(row["paid_vat_amount"]),
                float(row["paid_accountant_amount"]),
                float(row["paid_savings_amount"]),
                float(row["paid_tag_amount"]),
            )

            if existing:
                connection.execute(
                    """
                    UPDATE invoices
                    SET client = ?, description = ?, net_amount = ?, vat_rate = ?, vat_amount = ?,
                        tag_amount = ?, accountant_amount = ?, total_amount = ?,
                        deposit_amount = ?, savings_amount = ?, deposit_manuel_amount = ?,
                        paid_vat_amount = ?, paid_accountant_amount = ?, paid_savings_amount = ?,
                        paid_tag_amount = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (
                        client_name,
                        description,
                        float(row["net_amount"]),
                        vat_rate,
                        float(row["vat_amount"]) or totals["vat_amount"],
                        float(row["tag_amount"]),
                        float(row["accountant_amount"]),
                        float(row["total_amount"]) or totals["total_amount"],
                        float(row["deposit_amount"]),
                        float(row["savings_amount"]),
                        float(row["deposit_manuel_amount"]),
                        float(row["paid_vat_amount"]),
                        float(row["paid_accountant_amount"]),
                        float(row["paid_savings_amount"]),
                        float(row["paid_tag_amount"]),
                        int(existing["id"]),
                    ),
                )
                updated_invoices += 1
            else:
                connection.execute(
                    """
                    INSERT INTO invoices(
                        invoice_number, invoice_date, client, description,
                        net_amount, vat_rate, vat_amount, tag_amount,
                        accountant_amount, total_amount, deposit_amount, savings_amount,
                        deposit_manuel_amount, paid_vat_amount, paid_accountant_amount,
                        paid_savings_amount, paid_tag_amount, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    """,
                    payload,
                )
                created_invoices += 1

        for row in payment_rows:
            existing = connection.execute(
                "SELECT id FROM sii_reconciliations WHERE month = ?",
                (row["month"],),
            ).fetchone()
            if existing:
                connection.execute(
                    """
                    UPDATE sii_reconciliations
                    SET sii_vat_amount = ?, actual_tag_paid = ?, actual_accountant_paid = ?,
                        actual_savings_paid = ?,
                        observation = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (
                        float(row["sii_vat_amount"]),
                        float(row["actual_tag_paid"]),
                        float(row["actual_accountant_paid"]),
                        float(row.get("actual_savings_paid", 0)),
                        f"Importado desde {workbook_path.name}",
                        int(existing["id"]),
                    ),
                )
                updated_payments += 1
            else:
                connection.execute(
                    """
                    INSERT INTO sii_reconciliations(
                        month, sii_vat_amount, actual_tag_paid, actual_accountant_paid,
                        actual_savings_paid, observation, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    """,
                    (
                        row["month"],
                        float(row["sii_vat_amount"]),
                        float(row["actual_tag_paid"]),
                        float(row["actual_accountant_paid"]),
                        float(row.get("actual_savings_paid", 0)),
                        f"Importado desde {workbook_path.name}",
                    ),
                )
                created_payments += 1

        connection.commit()

    return created_invoices, updated_invoices, created_payments, updated_payments


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa un Excel real a Facturion")
    parser.add_argument("workbook", help="Ruta del Excel a importar")
    parser.add_argument("--client", default="Manuel", help="Nombre de cliente a usar en las facturas importadas")
    parser.add_argument("--cleanup", action="store_true", help="Limpia facturas y pagos antes de importar")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    created_invoices, updated_invoices, created_payments, updated_payments = import_workbook(
        workbook_path=workbook_path,
        client_name=args.client.strip() or "Manuel",
        cleanup=args.cleanup,
    )

    print(f"Facturas creadas: {created_invoices}")
    print(f"Facturas actualizadas: {updated_invoices}")
    print(f"Pagos creados: {created_payments}")
    print(f"Pagos actualizados: {updated_payments}")
    print("Importacion finalizada.")


if __name__ == "__main__":
    main()
